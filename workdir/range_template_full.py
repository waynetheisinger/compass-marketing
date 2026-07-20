import os, warnings, json
warnings.filterwarnings("ignore")
import openpyxl

DIR = "range/templates"

def dump(path):
    print("\n" + "=" * 90)
    print("FILE:", os.path.basename(path))
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    cols = list(wb["Columns"].iter_rows(values_only=True))
    header = cols[0]
    breadcrumb = header[4]
    print("CATEGORY BREADCRUMB:", breadcrumb)
    print(f"\n{len(cols)-1} columns (code | marker | label):")
    by_marker = {"REQUIRED": [], "RECOMMENDED": [], "OPTIONAL": [], None: []}
    for r in cols[1:]:
        code, label, desc, ex = r[0], r[1], r[2], r[3]
        marker = r[4] if len(r) > 4 else None
        by_marker.setdefault(marker, []).append(code)
        if marker == "REQUIRED":
            print(f"   [REQ]  {code:36} {label}")
    print("\n  marker tally:", {k: len(v) for k, v in by_marker.items()})
    print("  RECOMMENDED:", by_marker["RECOMMENDED"])

    # value lists from ReferenceData
    ref = list(wb["ReferenceData"].iter_rows(values_only=True))
    rhdr = ref[0]
    print(f"\n  value-list columns ({sum(1 for c in rhdr if c)}):")
    for ci, code in enumerate(rhdr):
        if not code:
            continue
        vals = [str(r[ci]) for r in ref[1:] if ci < len(r) and r[ci] not in (None, "")]
        preview = vals[:8]
        print(f"    {code:34} ({len(vals)})  {preview}")
    wb.close()

for f in ["Chainsaws.xlsx", "Cordless_Lawn_Mowers.xlsx", "Petrol_Lawn_Mowers.xlsx"]:
    dump(os.path.join(DIR, f))
