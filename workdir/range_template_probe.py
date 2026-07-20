import os, sys, warnings
warnings.filterwarnings("ignore")
import openpyxl

DIR = "range/templates"
FILES = ["Cordless_Lawn_Mowers.xlsx", "Chainsaws.xlsx", "Lawn_Mower_Accessories.xlsx"]

def show(path):
    print("\n" + "=" * 90)
    print("FILE:", path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print("SHEETS:", wb.sheetnames)
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        print(f"\n--- sheet '{sn}'  ({len(rows)} rows x {ws.max_column} cols) ---")
        for i, r in enumerate(rows[:6]):
            # trim long cells
            cells = [("" if c is None else str(c))[:34] for c in r]
            print(f"  r{i}: {cells}")
        if len(rows) > 6:
            print(f"  ... ({len(rows)-6} more rows)")
    wb.close()

for f in FILES:
    show(os.path.join(DIR, f))
