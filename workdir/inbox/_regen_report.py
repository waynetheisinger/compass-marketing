import zipfile, re, os
import openpyxl

src = 'workdir/inbox/EN_ES_Battery_Sellers.xlsx'            # fixed template
orig = 'workdir/inbox/EN_ES_Battery_Sellers.ORIGINAL.xlsx'  # pristine, for diff
out = 'reports/battery_exemption/SPECTRUM_battery_exemption_Amazon_kits_2026-06-02.xlsx'

def esc(s):
    return s.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

ANCHORS = [13, 17, 21, 25, 29]
rows = [
 (13, 'B0GYLMPTDM', 'SPECTRUM SBS460CLM 40V Cordless Self-Propelled Lawn Mower 46cm Kit',
  '46cm self-propelled lawn mower, 2 x 40V 4.0Ah lithium-ion battery packs (160Wh each, 10 cells each), 1 x dual-port fast charger', '101 - 300 WH'),
 (17, 'B0GYLNW2D5', 'SPECTRUM SBS220CHM 40V Cordless Lawn Mower 22cm Kit',
  '22cm lawn mower, 1 x 40V 2.0Ah lithium-ion battery pack (80Wh, 10 cells), 1 x charger', 'WH <= 100'),
 (21, 'B0GYLMG74Q', 'SPECTRUM SBS480CBV 40V Cordless Leaf Blower Vacuum Kit',
  '3-in-1 leaf blower vacuum, 1 x 40V 4.0Ah lithium-ion battery pack (160Wh, 10 cells), 1 x charger', '101 - 300 WH'),
 (25, 'B0GYLGNVYR', 'SPECTRUM SBS560CHT 40V Cordless Hedge Trimmer 45cm Kit',
  '45cm hedge trimmer, 1 x 40V 2.0Ah lithium-ion battery pack (80Wh, 10 cells), 1 x charger', 'WH <= 100'),
 (29, 'B0GYLGXC6J', 'SPECTRUM SBS240CPHT 40V Cordless Pole Hedge Trimmer Kit',
  '2.4m pole hedge trimmer, 1 x 40V 4.0Ah lithium-ion battery pack (160Wh, 10 cells), 1 x charger', '101 - 300 WH'),
]
cellvals = {}
for r, asin, title, box, whb in rows:
    cellvals[f'I{r}'] = asin
    cellvals[f'J{r}'] = title
    cellvals[f'K{r}'] = box
    cellvals[f'L{r}'] = 'Yes'
    cellvals[f'M{r}'] = 'Lithium_Ion'
    cellvals[f'N{r}'] = 'With equipment'
    cellvals[f'O{r}'] = 'Multiple_cells'
    cellvals[f'P{r}'] = whb
cellvals['D25'] = 'Wayne'
cellvals['G25'] = 'Theisinger'


def set_value_cell(xml, ref, val):
    """Fill an empty self-closed cell, preserving its style index."""
    pat = re.compile(rf'<c r="{ref}"((?:\s+[a-z]+="[^"]*")*?)\s*/>')
    m = pat.search(xml)
    if not m:
        return xml, False
    attrs = re.sub(r'\s+t="[^"]*"', '', m.group(1))
    repl = f'<c r="{ref}"{attrs} t="inlineStr"><is><t xml:space="preserve">{esc(val)}</t></is></c>'
    return xml[:m.start()] + repl + xml[m.end():], True


def set_cached(xml, ref, cached):
    """Overwrite the cached <v> of a formula cell (keeps the <f> formula intact)."""
    m = re.search(rf'(<c r="{ref}"[^>]*?>)(.*?)(</c>)', xml, re.S)
    if not m:
        return xml, False
    head, body, tail = m.groups()
    body2 = re.sub(r'<v[^>]*/>|<v>.*?</v>', f'<v>{esc(cached)}</v>', body, count=1)
    return xml[:m.start()] + head + body2 + tail + xml[m.end():], True


with zipfile.ZipFile(src) as zin:
    xml = zin.read('xl/worksheets/sheet1.xml').decode()
    wbxml = zin.read('xl/workbook.xml').decode()

# 1) inject the 42 data values
missing = []
for ref, val in cellvals.items():
    xml, ok = set_value_cell(xml, ref, val)
    if not ok:
        missing.append(ref)
print('cells injected:', len(cellvals) - len(missing), '| missing:', missing)

# 2) bake correct cached formula results (so it's right even without a recalc)
baked = []
for ref, cv in [('D24', ''), ('G24', ''), ('X13', '0')]:
    xml, ok = set_cached(xml, ref, cv); baked.append((ref, ok))
for r in ANCHORS:
    for col in ('W', 'R'):
        xml, ok = set_cached(xml, f'{col}{r}', 'Complete'); baked.append((f'{col}{r}', ok))
print('caches baked:', all(ok for _, ok in baked), '| failures:', [k for k, ok in baked if not ok])

# 3) force a full recalc on open (belt-and-suspenders with the baked caches)
if 'fullCalcOnLoad' not in wbxml:
    wbxml = re.sub(r'<calcPr ', '<calcPr fullCalcOnLoad="1" ', wbxml, count=1)
print('fullCalcOnLoad set:', 'fullCalcOnLoad="1"' in wbxml)

os.makedirs(os.path.dirname(out), exist_ok=True)
with zipfile.ZipFile(src) as zin, zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as zout:
    for it in zin.infolist():
        data = zin.read(it.filename)
        if it.filename == 'xl/worksheets/sheet1.xml':
            data = xml.encode()
        elif it.filename == 'xl/workbook.xml':
            data = wbxml.encode()
        zout.writestr(it, data)
print('WROTE', out)

# ---- verify ----
zo = zipfile.ZipFile(orig); zn = zipfile.ZipFile(out)
print('changed entries vs original:', [n for n in zo.namelist() if zo.read(n) != zn.read(n)])
s1 = zn.read('xl/worksheets/sheet1.xml').decode()
print('x14 dataValidation present:', 'x14:dataValidation' in s1)
print('sheetProtection present:', '<sheetProtection' in s1)
for ref in ['D24', 'G24', 'X13', 'R13', 'W13', 'R29']:
    m = re.search(rf'<c r="{ref}"[^>]*?>.*?</c>', s1, re.S)
    v = re.search(r'<v[^>]*>(.*?)</v>|<v[^>]*/>', m.group(0))
    print(f'  {ref} cached = {v.group(1) if v and v.group(1) is not None else "(empty)"!r}')

wb = openpyxl.load_workbook(out)
ws = wb['Battery exemption sheet']
fs = openpyxl.load_workbook(out, data_only=True)['Formula']
table = {}
for row in fs.iter_rows(min_col=8, max_col=15, values_only=True):
    if row[0] is not None:
        table[str(row[0])] = row[7]
print('\nrow | ASIN | -> computed status')
for r in ANCHORS:
    L, M, N, O, P = (ws.cell(r, c).value for c in (12, 13, 14, 15, 16))
    Q = ws.cell(r, 17).value or ''
    print(f'{r} | {ws.cell(r,9).value} | -> {table.get(f"{L}{M}{N}{O}{P}{Q}","NOT FOUND")}')
print('names:', ws['D25'].value, ws['G25'].value)
