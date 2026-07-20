import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

prods=json.load(open("workdir/dims_products.json"))
rows=list(prods.values())
# sort: multi-channel first, then title
rows.sort(key=lambda p:(0 if len(p["channels"])>1 else 1, p["channels"][0] if p["channels"] else "z", (p["title"] or "zz").lower()))

wb=Workbook(); ws=wb.active; ws.title="Product Dimensions"

TITLE_FILL=PatternFill("solid", fgColor="1F4E5F")
HDR_FILL  =PatternFill("solid", fgColor="2E75B6")
GRP1      =PatternFill("solid", fgColor="D9E1F2")  # product dims header
GRP2      =PatternFill("solid", fgColor="FCE4D6")  # shipping dims header
FILLIN    =PatternFill("solid", fgColor="FFF2CC")  # blank -> staff fill (amber)
PREFILL   =PatternFill("solid", fgColor="E2EFDA")  # from Amazon -> verify (green)
white=Font(color="FFFFFF", bold=True); bold=Font(bold=True)
thin=Side(style="thin", color="BFBFBF"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
wrap=Alignment(wrap_text=True, vertical="top"); ctr=Alignment(horizontal="center", vertical="center", wrap_text=True)

cols=[
 ("EAN",14),("Shopify SKU",16),("Product Title",44),("Marketplaces",14),
 ("Amazon SKU(s)",18),("B&Q SKU(s)",14),
 ("Product\nLength cm",10),("Product\nWidth cm",10),("Product\nHeight cm",10),("Product\nWeight kg",10),
 ("Shipping Box\nLength cm",11),("Shipping Box\nWidth cm",11),("Shipping Box\nHeight cm",11),("Shipping\nWeight kg",11),
 ("Notes",26),
]
NC=len(cols)

# Title + instructions
ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=NC)
c=ws.cell(1,1,"MowDirect — Product & Shipping Dimensions to Collect"); c.font=Font(color="FFFFFF",bold=True,size=14); c.fill=TITLE_FILL; c.alignment=Alignment(vertical="center")
ws.row_dimensions[1].height=26
ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=NC)
c=ws.cell(2,1,"Every product listed on a marketplace needs BOTH product dimensions (required by The Range) and shipping-box dimensions + weight (required for Amazon FBA and marketplace Click & Collect).  "
              "Amber cells = please measure & fill in.  Green cells = pulled from Amazon, please double-check.  Units: centimetres (cm) and kilograms (kg).")
c.font=Font(italic=True, size=10); c.alignment=wrap; ws.row_dimensions[2].height=42

# group header row (row 3)
for r in (3,):
    for ci,(name,_) in enumerate(cols, start=1):
        ws.cell(r,ci)
ws.merge_cells(start_row=3,start_column=7,end_row=3,end_column=10)
g=ws.cell(3,7,"PRODUCT DIMENSIONS  (The Range)"); g.fill=GRP1; g.font=bold; g.alignment=ctr; g.border=border
ws.merge_cells(start_row=3,start_column=11,end_row=3,end_column=14)
g=ws.cell(3,11,"SHIPPING / PACKAGE  (FBA + Click & Collect)"); g.fill=GRP2; g.font=bold; g.alignment=ctr; g.border=border

# column header row (row 4)
HR=4
for ci,(name,w) in enumerate(cols, start=1):
    cc=ws.cell(HR,ci,name); cc.fill=HDR_FILL; cc.font=white; cc.alignment=ctr; cc.border=border
    ws.column_dimensions[get_column_letter(ci)].width=w
ws.row_dimensions[HR].height=30

dimkeys=[("p_l"),("p_w"),("p_h"),("p_wt"),("s_l"),("s_w"),("s_h"),("s_wt")]
r=HR+1
for p in rows:
    vals=[
      p["ean"], p["shopify_sku"], p["title"], " + ".join(p["channels"]),
      ", ".join(dict.fromkeys(p["amz_skus"])), ", ".join(dict.fromkeys(p["bq_skus"])),
    ]
    for ci,v in enumerate(vals, start=1):
        cc=ws.cell(r,ci,v); cc.border=border; cc.alignment=wrap if ci==3 else Alignment(vertical="top")
    for j,k in enumerate(dimkeys):
        ci=7+j; val=p.get(k)
        cc=ws.cell(r,ci, val if val is not None else None)
        cc.border=border; cc.alignment=ctr
        cc.fill = PREFILL if val is not None else FILLIN
    ws.cell(r,NC).border=border  # notes
    r+=1

ws.freeze_panes="D5"
ws.auto_filter.ref=f"A{HR}:{get_column_letter(NC)}{r-1}"

out="reports/product_dimensions_to_collect.xlsx"
wb.save(out)
print("wrote", out, "with", r-1-HR, "product rows")
