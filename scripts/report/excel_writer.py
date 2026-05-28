"""
Excel spreadsheet builder for the monthly spend report.

Produces a 6-tab .xlsx workbook:

  Tab 1 — Summary
  Tab 2 — Marketplace Fees & Commissions
  Tab 3 — FBA Returns & Removals
  Tab 4 — Cancellations
  Tab 5 — Ad Spend
  Tab 6 — Raw Data (sample transaction rows)

The legacy "FBA Cost of Sales" tab and the matching Summary-tab section
were removed — funder report no longer itemises FBA operational cost.
A placeholder "Commission paid to Wayne Theisinger" line (£0 default) now
occupies that slot on the Summary tab; set via `wayne_commission` on
`transforms.build_summary`.

Live formulas ("real spreadsheet")
----------------------------------
Derived cells (subtotals, grand totals, every percentage / ratio, and the
Summary tab's restated channel & ad-spend figures) are written as **Excel
formulas**, not pre-computed values — so an auditor can click any total and
see the calculation, and changing a detail-tab figure recalculates the
Summary. Leaf cells (per-fee amounts, unit counts, per-campaign spend) stay
as plain values because they come straight from the source APIs and have
nothing to derive from.

Single source of truth: the detail tabs are built first and register the
coordinates of their key cells in an `anchors` dict; the Summary tab is
built last and cross-references those cells (e.g.
``='Marketplace Fees & Commissions'!C25``). The sheets are then reordered so
Summary appears first.

openpyxl writes the formula text but does not compute it, so the workbook is
flagged ``fullCalcOnLoad`` — Excel / Google Sheets / LibreOffice recalculate
every formula the moment the file is opened. (Anything that reads the *cached*
value without opening in a spreadsheet app — e.g. macOS Quick Look — would
see blanks; the agreed consumer is humans opening in Excel/Sheets.)

All division formulas are wrapped in ``IFERROR(…, 0)`` so a zero denominator
(e.g. a NOT CONNECTED channel with £0 net) renders 0 rather than #DIV/0!.

Usage:
    from scripts.report.excel_writer import build_workbook
    wb = build_workbook(report_data, month_label="March 2026")
    wb.save("reports/marketing_spend_2026-03.xlsx")
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Sheet titles (used for cross-sheet references — keep in one place)
# ---------------------------------------------------------------------------
SHEET_SUMMARY     = "Summary"
SHEET_MARKETPLACE = "Marketplace Fees & Commissions"
SHEET_FBA         = "FBA Returns & Removals"
SHEET_CANCEL      = "Cancellations"
SHEET_ADSPEND     = "Ad Spend"
SHEET_RAW         = "Raw Data"

# Fallback commission rate if report_data["summary"] doesn't carry one
# (mirrors WAYNE_COMMISSION_RATE in monthly_report.py).
_DEFAULT_COMMISSION_RATE = 0.04

# ---------------------------------------------------------------------------
# Colour palette (matches mock_report.py)
# ---------------------------------------------------------------------------
DARK_GREEN  = "1B5E20"
MID_GREEN   = "2E7D32"
LIGHT_GREEN = "C8E6C9"
PALE_GREEN  = "E8F5E9"
DARK_BLUE   = "0D47A1"
MID_BLUE    = "1565C0"
LIGHT_BLUE  = "BBDEFB"
PALE_BLUE   = "E3F2FD"
AMBER       = "FFF8E1"
WHITE       = "FFFFFF"
DARK_TEXT   = "212121"
MUTED_TEXT  = "757575"

GBP = "£#,##0.00"
PCT = "0.0%"
INT = "#,##0"


# ---------------------------------------------------------------------------
# Cell-reference helpers (for building formulas)
# ---------------------------------------------------------------------------

def _a1(col: int, row: int) -> str:
    """Return an A1-style reference for a (column, row) pair, e.g. (3, 4) -> C4."""
    return f"{get_column_letter(col)}{row}"


def _xref(sheet_title: str, col: int, row: int) -> str:
    """
    Cross-sheet cell reference as a standalone formula, sheet name quoted
    (handles spaces / '&'), e.g. ``='Ad Spend'!C40``.
    """
    return f"='{sheet_title}'!{_a1(col, row)}"


def _sum(col: int, row_start: int, row_end: int) -> str:
    """=SUM over a contiguous single-column range."""
    return f"=SUM({_a1(col, row_start)}:{_a1(col, row_end)})"


def _ratio(num_col: int, num_row: int, den_col: int, den_row: int) -> str:
    """=IFERROR(num/den, 0) — safe ratio that renders 0 on a zero denominator."""
    return f"=IFERROR({_a1(num_col, num_row)}/{_a1(den_col, den_row)},0)"


# ---------------------------------------------------------------------------
# Low-level style helpers
# ---------------------------------------------------------------------------

def _fill(hex_colour: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_colour)


def _font(bold=False, size=10, colour=DARK_TEXT, italic=False) -> Font:
    return Font(bold=bold, size=size, color=colour, italic=italic)


def _border_bottom(colour="E0E0E0") -> Border:
    return Border(bottom=Side(style="thin", color=colour))


def _set_col_widths(ws, widths: list[float]) -> None:
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


# ---------------------------------------------------------------------------
# Higher-level cell writers
# ---------------------------------------------------------------------------

def _title(ws, text: str, ncols: int, bg=DARK_GREEN) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.row_dimensions[1].height = 34
    c = ws["A1"]
    c.value = text
    c.fill = _fill(bg)
    c.font = Font(bold=True, size=13, color=WHITE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def _header_row(ws, row: int, values: list, bg=DARK_GREEN, fg=WHITE) -> None:
    ws.row_dimensions[row].height = 22
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = _fill(bg)
        c.font = Font(bold=True, size=10, color=fg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _cell(ws, row: int, col: int, value, fmt=None, bold=False,
          bg=WHITE, align="right", fg=DARK_TEXT, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    if fmt:
        c.number_format = fmt
    c.font = Font(bold=bold, size=10, color=fg, italic=italic)
    c.fill = _fill(bg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = _border_bottom()
    return c


def _section_heading(ws, row: int, text: str, ncols: int, bg=MID_GREEN) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.row_dimensions[row].height = 20
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _fill(bg)
    c.font = Font(bold=True, size=10, color=WHITE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def _subtotal_row(ws, row: int, ncols: int, label: str,
                  col_values: dict, bg=LIGHT_GREEN, fg=DARK_TEXT) -> None:
    for col in range(1, ncols + 1):
        ws.cell(row=row, column=col).fill = _fill(bg)
    _cell(ws, row, 1, label, align="right", bg=bg, bold=True, fg=fg)
    for col, (val, fmt) in col_values.items():
        _cell(ws, row, col, val, fmt=fmt, bg=bg, bold=True, fg=fg)


def _grand_total_row(ws, row: int, ncols: int, label: str,
                     col_values: dict, bg=DARK_GREEN) -> None:
    ws.row_dimensions[row].height = 18
    for col in range(1, ncols + 1):
        ws.cell(row=row, column=col).fill = _fill(bg)
        ws.cell(row=row, column=col).font = Font(bold=True, size=10, color=WHITE)
    _cell(ws, row, 1, label, align="left", bg=bg, bold=True, fg=WHITE)
    for col, (val, fmt) in col_values.items():
        c = _cell(ws, row, col, val, fmt=fmt, bg=bg, bold=True, fg=WHITE)
        c.font = Font(bold=True, size=10, color=WHITE)


def _note_row(ws, row: int, ncols: int, text: str,
              bg=PALE_BLUE, fg=DARK_BLUE, height: float = 20, wrap: bool = False) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.row_dimensions[row].height = height
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _fill(bg)
    c.font = Font(size=9, color=fg, italic=True)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=wrap)


def _coverage_cell(ws, row: int, col: int, note: str | None) -> None:
    """Render a coverage/status cell — green COMPLETE or amber NOT CONNECTED."""
    if note is None:
        _cell(ws, row, col, "COMPLETE", align="center",
              bg="E8F5E9", fg=MID_GREEN, bold=True)
    elif note.startswith("PARTIAL"):
        _cell(ws, row, col, "PARTIAL", align="center",
              bg=AMBER, fg="E65100", bold=True)
    else:
        _cell(ws, row, col, "NOT CONNECTED", align="center",
              bg="FFF3E0", fg="BF360C", bold=True)


def _mock_banner(ws, row: int, ncols: int) -> None:
    """Amber banner shown when running in dry-run / mock mode."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.row_dimensions[row].height = 22
    c = ws.cell(row=row, column=1,
                value="⚠  MOCK / DRY-RUN DATA — replace with live API credentials before sharing with funders.")
    c.fill = _fill(AMBER)
    c.font = Font(bold=True, size=9, color="E65100")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)


# ---------------------------------------------------------------------------
# Tab 2 — Marketplace Fees & Commissions  (built first; source of truth)
# ---------------------------------------------------------------------------

def _build_marketplace(wb: Workbook, data: dict, month_label: str, anchors: dict) -> None:
    ws = wb.create_sheet(SHEET_MARKETPLACE)
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [24, 36, 14, 16, 14, 26, 16])
    _title(ws, f"Marketplace Commissions & Fees  |  {month_label}", 7)
    ws.row_dimensions[2].height = 4

    _header_row(ws, 3, [
        "Channel", "Fee Type", "Amount (£)", "Net Channel Revenue (£)",
        "Fee as % of Net Revenue", "Data Source", "Coverage",
    ])

    summary = data["summary"]
    net     = summary["net"]
    r = 4

    channel_subtotal_rows: list[int] = []   # one per channel (for grand total)

    for ch in data["channels"]:
        ch_fees = ch["total_fees"]
        # Suppress zero-amount fee types (e.g. Amazon's FixedClosingFee,
        # GiftwrapCommission — SP-API returns them even when £0). If a
        # channel has *only* zero rows (typically a NOT CONNECTED channel),
        # keep the first row so the channel header + coverage cell still
        # appear on the tab.
        nonzero_fees = [r for r in ch["fee_rows"] if float(r.get("amount", 0) or 0) != 0]
        # Always render at least one row per channel: a channel with no fees at
        # all this period (e.g. Amazon before settlements post mid-month) still
        # needs a row so its net cell is written and flows into the grand total,
        # and so the subtotal's SUM range is valid rather than backwards.
        fee_rows = (nonzero_fees or ch["fee_rows"][:1]
                    or [{"label": "No fees this period", "amount": 0.0}])

        first_fee_row = r
        net_cell_row  = r            # net is written on the channel's first row
        for j, row_data in enumerate(fee_rows):
            bg = PALE_GREEN if r % 2 == 0 else WHITE
            _cell(ws, r, 1, ch["name"] if j == 0 else "",
                  align="left", bg=bg, bold=(j == 0))
            _cell(ws, r, 2, row_data["label"], align="left", bg=bg)
            _cell(ws, r, 3, row_data["amount"], fmt=GBP, bg=bg)        # leaf data
            _cell(ws, r, 4, ch["net"] if j == 0 else None,            # leaf data
                  fmt=GBP if j == 0 else "@", bg=bg)
            # Fee % = this fee amount / channel net  (formula)
            _cell(ws, r, 5, _ratio(3, r, 4, net_cell_row), fmt=PCT, bg=bg)
            _cell(ws, r, 6, ch["source"] if j == 0 else "",
                  align="left", bg=bg, fg=MUTED_TEXT, italic=True)
            if j == 0:
                _coverage_cell(ws, r, 7, ch.get("note"))
            else:
                _cell(ws, r, 7, "", bg=bg)
            r += 1
        last_fee_row = r - 1

        # Channel subtotal — fees SUM the rows above; net references the
        # channel's single net cell; fee % divides the two.
        _subtotal_row(ws, r, 7, f"{ch['name']} subtotal", {
            3: (_sum(3, first_fee_row, last_fee_row), GBP),
            4: (f"={_a1(4, net_cell_row)}", GBP),
            5: (_ratio(3, r, 4, r), PCT),
        })
        anchors["mp"][ch["name"]] = {"fees_row": r, "net_row": r}
        channel_subtotal_rows.append(r)
        r += 1

    # Grand total — sum the per-channel subtotal cells (matches the visual
    # structure: "the total is the sum of the subtotals").
    gt_row = r
    fees_gt = "=" + "+".join(_a1(3, rr) for rr in channel_subtotal_rows) if channel_subtotal_rows else 0
    net_gt  = "=" + "+".join(_a1(4, rr) for rr in channel_subtotal_rows) if channel_subtotal_rows else 0
    _grand_total_row(ws, gt_row, 7, "GRAND TOTAL — MARKETPLACE FEES", {
        3: (fees_gt, GBP),
        4: (net_gt, GBP),
        5: (_ratio(3, gt_row, 4, gt_row), PCT),
    })
    anchors["mp_grand_fees_row"] = gt_row
    anchors["mp_grand_net_row"]  = gt_row
    r += 2

    # Deliberate margin note — a real finding, not to be massaged out of the data.
    _note_row(ws, r, 7,
              "ℹ  Margin note — Amazon's referral fee is ~15% of GROSS (inc-VAT) sales, "
              "which works out to a HIGHER share of NET (ex-VAT) revenue — ~18-19%, as "
              "shown above. This is a real effect to be aware of: Amazon's effective cost "
              "against net revenue genuinely exceeds the 15% headline rate. It is surfaced "
              "here deliberately rather than masked. (Fees are already shown ex-VAT — the "
              "reclaimable 20% VAT Amazon charges on its fees has been removed.)",
              bg=AMBER, fg="E65100", height=58, wrap=True)
    r += 2

    if data.get("is_mock"):
        _mock_banner(ws, r, 7)


# ---------------------------------------------------------------------------
# Tab 3 — FBA Returns & Removals
# ---------------------------------------------------------------------------

def _build_fba_returns(wb: Workbook, data: dict, month_label: str, anchors: dict) -> None:
    """
    Customer returns by disposition, removal shipments (return-to-seller vs
    disposal vs liquidation), unfulfillable inventory snapshot, and the £ cost
    of removal/disposal activity. Totals and percentages are formulas.
    """
    ws = wb.create_sheet(SHEET_FBA)
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [40, 14, 14, 16, 38])
    _title(ws, f"Amazon FBA — Returns, Removals & Stranded Stock  |  {month_label}",
           5, bg=DARK_BLUE)
    ws.row_dimensions[2].height = 4

    note = data.get("fba_returns_note")
    returns_data = data.get("fba_returns") or {}

    r = 3
    if note and not returns_data:
        _note_row(ws, r, 5, note, bg="FFF3E0", fg="BF360C")
        if data.get("is_mock"):
            _mock_banner(ws, r + 2, 5)
        return

    if note:  # PARTIAL — render data we have but flag what failed
        _note_row(ws, r, 5, note, bg=AMBER, fg="E65100")
        r += 2

    cust    = returns_data.get("customer_returns_summary", {})
    removal = returns_data.get("removal_shipments_summary", {})
    snap    = returns_data.get("inventory_snapshot_summary", {})
    fees    = returns_data.get("removal_fee_totals", {})

    # ── Customer returns by disposition ────────────────────────────────
    _section_heading(ws, r, "  Customer Returns Received by Amazon", 5, bg=MID_BLUE)
    r += 1
    _header_row(ws, r,
                ["Disposition", "Units", "Lines", "% of Returns", "Notes"],
                bg=DARK_BLUE)
    r += 1

    by_disposition = cust.get("by_disposition", [])
    if not by_disposition:
        _cell(ws, r, 1, "No customer returns received this month.",
              align="left", bg=PALE_BLUE, italic=True, fg=MUTED_TEXT)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1
    else:
        first_disp_row = r
        # The TOTAL row lands directly after the disposition rows; % cells
        # reference its Units cell (col 2). Forward reference — fine in Excel.
        total_row = first_disp_row + len(by_disposition)
        for row_data in by_disposition:
            bg = PALE_BLUE if r % 2 == 0 else WHITE
            _cell(ws, r, 1, row_data["label"], align="left", bg=bg)
            _cell(ws, r, 2, row_data["units"], fmt=INT, bg=bg)        # leaf
            _cell(ws, r, 3, row_data["lines"], fmt=INT, bg=bg)        # leaf
            _cell(ws, r, 4, _ratio(2, r, 2, total_row), fmt=PCT, bg=bg)
            _cell(ws, r, 5, "", bg=bg)
            r += 1
        last_disp_row = r - 1

        _subtotal_row(ws, r, 5, "TOTAL CUSTOMER RETURNS", {
            2: (_sum(2, first_disp_row, last_disp_row), INT),
            3: (_sum(3, first_disp_row, last_disp_row), INT),
            4: (_ratio(2, r, 2, r), PCT),
        }, bg=LIGHT_BLUE, fg=DARK_BLUE)
        assert r == total_row, "customer-returns total row drifted"
        r += 1

        # Sellable vs unsellable highlight (units are aggregates → leaf data;
        # % references the total-units cell).
        sellable   = cust.get("sellable_units", 0)
        unsellable = cust.get("unsellable_units", 0)
        bg = PALE_BLUE
        _cell(ws, r, 1, "↳ Returned to active inventory", align="left", bg=bg, italic=True)
        _cell(ws, r, 2, sellable, fmt=INT, bg=bg)
        _cell(ws, r, 3, "", bg=bg)
        _cell(ws, r, 4, _ratio(2, r, 2, total_row), fmt=PCT, bg=bg)
        _cell(ws, r, 5, "Available for re-sale immediately.",
              align="left", bg=bg, fg=MUTED_TEXT, italic=True)
        r += 1
        _cell(ws, r, 1, "↳ Stranded as unfulfillable", align="left", bg=bg, italic=True)
        _cell(ws, r, 2, unsellable, fmt=INT, bg=bg)
        _cell(ws, r, 3, "", bg=bg)
        _cell(ws, r, 4, _ratio(2, r, 2, total_row), fmt=PCT, bg=bg)
        _cell(ws, r, 5, "Sit at Amazon until removed, disposed of, or aged out.",
              align="left", bg=bg, fg=MUTED_TEXT, italic=True)
        r += 2

    # ── Removal shipments by type ──────────────────────────────────────
    _section_heading(ws, r, "  Removal Shipments (units processed by Amazon this month)",
                     5, bg=MID_BLUE)
    r += 1
    _header_row(ws, r, ["Order Type", "Units", "Shipments", "", "Meaning"],
                bg=DARK_BLUE)
    r += 1

    type_meaning = {
        "Return":       "Shipped back to MowDirect.",
        "Disposal":     "Destroyed by Amazon at our request.",
        "Liquidations": "Sold to Amazon's liquidation partner.",
    }

    by_type = removal.get("by_type", [])
    if not by_type:
        _cell(ws, r, 1, "No removal shipments processed this month.",
              align="left", bg=PALE_BLUE, italic=True, fg=MUTED_TEXT)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1
    else:
        first_type_row = r
        for row_data in by_type:
            bg = PALE_BLUE if r % 2 == 0 else WHITE
            _cell(ws, r, 1, row_data["order_type"], align="left", bg=bg, bold=True)
            _cell(ws, r, 2, row_data["units"], fmt=INT, bg=bg)        # leaf
            _cell(ws, r, 3, row_data["lines"], fmt=INT, bg=bg)        # leaf
            _cell(ws, r, 4, "", bg=bg)
            _cell(ws, r, 5, type_meaning.get(row_data["order_type"], ""),
                  align="left", bg=bg, fg=MUTED_TEXT, italic=True)
            r += 1
        _subtotal_row(ws, r, 5, "TOTAL REMOVAL UNITS", {
            2: (_sum(2, first_type_row, r - 1), INT),
        }, bg=LIGHT_BLUE, fg=DARK_BLUE)
        r += 1

    r += 1

    # ── Removal / disposal fee costs ───────────────────────────────────
    _section_heading(ws, r, "  Removal & Disposal Costs (from Amazon Finances)", 5,
                     bg=MID_BLUE)
    r += 1
    _header_row(ws, r, ["Fee Type", "Amount (£)", "", "", "Notes"], bg=DARK_BLUE)
    r += 1

    fee_notes = {
        "FBA removal fees":       "Per-unit fee charged when removing stranded stock.",
        "FBA disposal fees":      "Per-unit fee charged for destruction.",
        "FBA removal shipping":   "Carrier costs for removal-back-to-seller shipments.",
        "Buyer return shipping":  "Free customer return shipping recovered from seller.",
    }

    if not fees:
        _cell(ws, r, 1, "No removal / disposal fees billed this month.",
              align="left", bg=PALE_BLUE, italic=True, fg=MUTED_TEXT)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1
    else:
        first_fee_row = r
        for label, amount in sorted(fees.items(), key=lambda x: -x[1]):
            bg = PALE_BLUE if r % 2 == 0 else WHITE
            _cell(ws, r, 1, label, align="left", bg=bg)
            _cell(ws, r, 2, amount, fmt=GBP, bg=bg)                   # leaf
            _cell(ws, r, 3, "", bg=bg)
            _cell(ws, r, 4, "", bg=bg)
            _cell(ws, r, 5, fee_notes.get(label, ""),
                  align="left", bg=bg, fg=MUTED_TEXT, italic=True)
            r += 1
        _subtotal_row(ws, r, 5, "TOTAL REMOVAL / DISPOSAL COST", {
            2: (_sum(2, first_fee_row, r - 1), GBP),
        }, bg=LIGHT_BLUE, fg=DARK_BLUE)
        r += 1

    r += 1

    # ── Current unfulfillable inventory snapshot ───────────────────────
    _section_heading(ws, r,
                     "  Current Inventory Snapshot — units sitting at Amazon today",
                     5, bg=MID_BLUE)
    r += 1
    _header_row(ws, r, ["Bucket", "Units", "", "", "Meaning"], bg=DARK_BLUE)
    r += 1

    snapshot_rows = [
        ("Fulfillable (sellable)",          snap.get("fulfillable", 0),
         "Active inventory available to ship to customers."),
        ("Unfulfillable",                   snap.get("unfulfillable", 0),
         "Stranded — could be removed or disposed."),
        ("Researching",                     snap.get("researching", 0),
         "Lost/damaged claims under investigation by Amazon."),
        ("Inbound working",                 snap.get("inbound_working", 0),
         "Shipment created in Seller Central but not yet sent."),
        ("Inbound shipped",                 snap.get("inbound_shipped", 0),
         "In transit from MowDirect/supplier to fulfilment centre."),
        ("Inbound receiving",               snap.get("inbound_receiving", 0),
         "Arrived at FC, being checked in."),
    ]
    unfulfillable_row = researching_row = None
    for label, units, meaning in snapshot_rows:
        bg = PALE_BLUE if r % 2 == 0 else WHITE
        is_pickup = label in {"Unfulfillable", "Researching"}
        _cell(ws, r, 1, label, align="left", bg=bg, bold=is_pickup)
        _cell(ws, r, 2, units, fmt=INT, bg=bg, bold=is_pickup)        # leaf
        _cell(ws, r, 3, "", bg=bg)
        _cell(ws, r, 4, "", bg=bg)
        _cell(ws, r, 5, meaning, align="left", bg=bg, fg=MUTED_TEXT, italic=True)
        if label == "Unfulfillable":
            unfulfillable_row = r
        elif label == "Researching":
            researching_row = r
        r += 1

    # Available to be picked up / disposed = Unfulfillable + Researching
    pickup_formula = f"={_a1(2, unfulfillable_row)}+{_a1(2, researching_row)}"
    _subtotal_row(ws, r, 5, "AVAILABLE TO BE PICKED UP / DISPOSED", {
        2: (pickup_formula, INT),
    }, bg=LIGHT_BLUE, fg=DARK_BLUE)
    r += 2

    sku_count = snap.get("sku_count", 0)
    _note_row(ws, r, 5,
              f"ℹ  Inventory snapshot covers {sku_count} active FBA SKUs at the "
              "time of report generation. Customer-returns and removal-shipment "
              "figures cover the reporting month only.")
    r += 2

    if data.get("is_mock"):
        _mock_banner(ws, r, 5)


# ---------------------------------------------------------------------------
# Tab 4 — Cancellations
# ---------------------------------------------------------------------------

def _build_cancellations(wb: Workbook, data: dict, month_label: str, anchors: dict) -> None:
    """
    Cancelled orders by channel, broken out by attribution where the
    platform exposes it (customer / seller). Direct channel APIs only —
    BaseLinker excluded to avoid double-counting against Amazon SP-API.
    Subtotals and the grand total are formulas.
    """
    ws = wb.create_sheet(SHEET_CANCEL)
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [22, 36, 12, 16, 38])
    _title(ws, f"Order Cancellations  |  {month_label}", 5)
    ws.row_dimensions[2].height = 4

    cancellations = data.get("cancellations") or {}
    notes         = data.get("cancellation_notes") or {}

    r = 3
    _header_row(ws, r, ["Channel", "Reason / Attribution",
                         "Orders", "Value (£)", "Notes / Coverage"])
    r += 1

    subtotal_order_rows: list[int] = []   # channel subtotal rows (col 3)
    subtotal_value_rows: list[int] = []   # channel subtotal rows (col 4)

    for channel_name, channel_summary in cancellations.items():
        if not channel_summary and channel_name not in notes:
            continue

        by_reason = (channel_summary or {}).get("by_reason", [])

        if not by_reason:
            bg = PALE_GREEN if r % 2 == 0 else WHITE
            _cell(ws, r, 1, channel_name, align="left", bg=bg, bold=True)
            note = notes.get(channel_name)
            if note:
                _cell(ws, r, 2, "—", align="center", bg=bg, fg=MUTED_TEXT)
                _cell(ws, r, 3, "—", align="center", bg=bg, fg=MUTED_TEXT)
                _cell(ws, r, 4, "—", align="center", bg=bg, fg=MUTED_TEXT)
                _coverage_cell(ws, r, 5, note)
            else:
                _cell(ws, r, 2, "No cancellations this month.",
                      align="left", bg=bg, italic=True, fg=MUTED_TEXT)
                _cell(ws, r, 3, 0, fmt=INT, bg=bg)
                _cell(ws, r, 4, 0, fmt=GBP, bg=bg)
                _cell(ws, r, 5, "", bg=bg)
            r += 1
            continue

        first_reason_row = r
        for j, row_data in enumerate(by_reason):
            bg = PALE_GREEN if r % 2 == 0 else WHITE
            _cell(ws, r, 1, channel_name if j == 0 else "",
                  align="left", bg=bg, bold=(j == 0))
            _cell(ws, r, 2, row_data["label"], align="left", bg=bg)
            _cell(ws, r, 3, row_data["orders"], fmt=INT, bg=bg)       # leaf
            _cell(ws, r, 4, row_data["value"],  fmt=GBP, bg=bg)       # leaf
            if j == 0:
                _coverage_cell(ws, r, 5, notes.get(channel_name))
            else:
                _cell(ws, r, 5, "", bg=bg)
            r += 1

        _subtotal_row(ws, r, 5, f"{channel_name} subtotal", {
            3: (_sum(3, first_reason_row, r - 1), INT),
            4: (_sum(4, first_reason_row, r - 1), GBP),
        })
        subtotal_order_rows.append(r)
        subtotal_value_rows.append(r)
        r += 1

    orders_gt = ("=" + "+".join(_a1(3, rr) for rr in subtotal_order_rows)
                 if subtotal_order_rows else 0)
    value_gt  = ("=" + "+".join(_a1(4, rr) for rr in subtotal_value_rows)
                 if subtotal_value_rows else 0)
    _grand_total_row(ws, r, 5, "TOTAL CANCELLATIONS", {
        3: (orders_gt, INT),
        4: (value_gt,  GBP),
    })
    r += 2

    _note_row(ws, r, 5,
              "ℹ  Direct channel APIs only — Shopify (cancelReason), Amazon SP-API "
              "(IsBuyerRequestedCancellation), Mirakl/B&Q (state). ManoMano, OnBuy "
              "and eBay cancellations are not yet covered.")
    r += 2

    if data.get("is_mock"):
        _mock_banner(ws, r, 5)


# ---------------------------------------------------------------------------
# Tab 5 — Ad Spend  (built before Summary; registers per-platform anchors)
# ---------------------------------------------------------------------------

def _build_ad_spend(wb: Workbook, data: dict, month_label: str, anchors: dict) -> None:
    ws = wb.create_sheet(SHEET_ADSPEND)
    ws.sheet_view.showGridLines = False
    # Platform | Campaign | Spend | Conversions | Conv Value | Cost/Conv | ROAS | Impr | Clicks | CTR | Coverage
    _set_col_widths(ws, [18, 38, 12, 12, 14, 14, 10, 12, 10, 10, 18])
    _title(ws, f"Paid Ad Spend  |  {month_label}", 11)
    ws.row_dimensions[2].height = 4

    _header_row(ws, 3, [
        "Platform", "Campaign", "Spend (£)",
        "Conversions", "Conv. Value (£)", "Cost / Conv (£)", "ROAS",
        "Impressions", "Clicks", "CTR", "Coverage",
    ])

    ad_rows  = data.get("ad_spend_rows", [])
    # Filter: campaigns with spend > 0 only (Google Ads has hundreds of zero-spend campaigns)
    ad_rows = [r_ for r_ in ad_rows if (r_.get("spend") or 0) > 0]

    r = 4

    platforms: dict[str, list[dict]] = {}
    for row_data in ad_rows:
        p = row_data["platform"]
        platforms.setdefault(p, []).append(row_data)

    # Per-platform "total" cells we can reference for the grand total and for
    # the Summary cross-link. For a single-campaign platform this is the lone
    # campaign row; for a multi-campaign platform it's the subtotal row.
    platform_spend_rows: list[int] = []
    platform_conv_rows:  list[int] = []   # only platforms that have conversions
    platform_cval_rows:  list[int] = []

    for platform, rows in platforms.items():
        # Sort by spend desc within platform so the heavy hitters lead
        rows = sorted(rows, key=lambda x: -(x.get("spend") or 0))
        platform_total_conv = sum((r_.get("conversions") or 0) for r_ in rows)

        first_row = r
        for row_data in rows:
            bg = PALE_GREEN if r % 2 == 0 else WHITE
            conv  = row_data.get("conversions")
            conv_value = row_data.get("conversions_value")
            impr   = row_data.get("impressions")
            clicks = row_data.get("clicks")

            _cell(ws, r, 1, platform if r == first_row else "",
                  align="left", bg=bg, bold=(r == first_row))
            _cell(ws, r, 2, row_data["campaign_name"], align="left", bg=bg)
            _cell(ws, r, 3, row_data["spend"], fmt=GBP, bg=bg)        # leaf

            if conv is not None:
                _cell(ws, r, 4, round(float(conv), 2), fmt="0.00", bg=bg)   # leaf
            else:
                _cell(ws, r, 4, "—", align="center", bg=bg, fg=MUTED_TEXT)

            if conv_value is not None:
                _cell(ws, r, 5, round(float(conv_value), 2), fmt=GBP, bg=bg)  # leaf
            else:
                _cell(ws, r, 5, "—", align="center", bg=bg, fg=MUTED_TEXT)

            # Cost / Conv = spend / conversions
            if conv and float(conv) > 0:
                _cell(ws, r, 6, _ratio(3, r, 4, r), fmt=GBP, bg=bg)
            else:
                _cell(ws, r, 6, "—", align="center", bg=bg, fg=MUTED_TEXT)

            # ROAS = conv value / spend
            if conv_value and row_data["spend"] > 0:
                _cell(ws, r, 7, _ratio(5, r, 3, r), fmt="0.00", bg=bg)
            else:
                _cell(ws, r, 7, "—", align="center", bg=bg, fg=MUTED_TEXT)

            _cell(ws, r, 8, impr,   fmt=INT if impr   else "@", bg=bg)   # leaf
            _cell(ws, r, 9, clicks, fmt=INT if clicks else "@", bg=bg)   # leaf
            # CTR = clicks / impressions
            if impr and clicks:
                _cell(ws, r, 10, _ratio(9, r, 8, r), fmt=PCT, bg=bg)
            else:
                _cell(ws, r, 10, "—", align="center", bg=bg, fg=MUTED_TEXT)
            _coverage_cell(ws, r, 11, data.get("ad_spend_notes", {}).get(platform))
            r += 1
        last_row = r - 1

        if len(rows) > 1:
            subtotal_cells: dict = {3: (_sum(3, first_row, last_row), GBP)}
            if platform_total_conv:
                subtotal_cells[4] = (_sum(4, first_row, last_row), "0.00")
                subtotal_cells[5] = (_sum(5, first_row, last_row), GBP)
                subtotal_cells[6] = (_ratio(3, r, 4, r), GBP)
                subtotal_cells[7] = (_ratio(5, r, 3, r), "0.00")
            _subtotal_row(ws, r, 11, f"{platform} subtotal", subtotal_cells)
            spend_total_row = r
            r += 1
        else:
            spend_total_row = first_row   # single campaign row is the total

        platform_spend_rows.append(spend_total_row)
        if platform_total_conv:
            platform_conv_rows.append(spend_total_row)
            platform_cval_rows.append(spend_total_row)
        anchors["ads_platform"][platform] = spend_total_row

    # Not-connected platforms
    for row_data in data.get("ad_spend_not_connected", []):
        bg = "FFF3E0"
        _cell(ws, r, 1, row_data["platform"], align="left", bg=bg, bold=True)
        for col in range(2, 11):
            _cell(ws, r, col, "—", align="center", bg=bg, fg=MUTED_TEXT)
        _cell(ws, r, 11, "NOT CONNECTED", align="center", bg=AMBER, fg="BF360C", bold=True)
        r += 1

    r += 1
    gt_row = r
    grand_cells: dict = {
        3: ("=" + "+".join(_a1(3, rr) for rr in platform_spend_rows) if platform_spend_rows else 0, GBP),
    }
    if platform_conv_rows:
        grand_cells[4] = ("=" + "+".join(_a1(4, rr) for rr in platform_conv_rows), "0.00")
        grand_cells[5] = ("=" + "+".join(_a1(5, rr) for rr in platform_cval_rows), GBP)
        grand_cells[6] = (_ratio(3, gt_row, 4, gt_row), GBP)
        grand_cells[7] = (_ratio(5, gt_row, 3, gt_row), "0.00")
    _grand_total_row(ws, gt_row, 11, "TOTAL AD SPEND", grand_cells)
    anchors["ads_grand_row"] = gt_row
    r += 2

    _note_row(ws, r, 11,
              "ℹ  Showing campaigns with spend only. Conversions / Conv. Value / "
              "Cost-per-Conv / ROAS reflect Google Ads conversion tracking; "
              "eBay Promoted Listings does not yet feed conversion data "
              "into this report.")
    r += 2

    if data.get("is_mock"):
        _mock_banner(ws, r, 11)


# ---------------------------------------------------------------------------
# Tab 1 — Summary  (built last; cross-references the detail tabs)
# ---------------------------------------------------------------------------

def _build_summary(wb: Workbook, data: dict, month_label: str, anchors: dict) -> None:
    ws = wb.create_sheet(SHEET_SUMMARY)
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [38, 18, 16, 44])
    _title(ws, f"MowDirect — Monthly Cost Report  |  {month_label}", 4)

    ws.row_dimensions[2].height = 4
    _header_row(ws, 3, ["Metric", "Amount (£)", "% of Net Revenue", "Notes"])

    summary = data["summary"]
    rate    = summary.get("commission_rate", _DEFAULT_COMMISSION_RATE)

    # Cross-sheet references into the detail tabs (the single source of truth)
    ref_mp_net   = _xref(SHEET_MARKETPLACE, 4, anchors["mp_grand_net_row"])
    ref_mp_fees  = _xref(SHEET_MARKETPLACE, 3, anchors["mp_grand_fees_row"])
    ref_ads      = _xref(SHEET_ADSPEND,     3, anchors["ads_grand_row"])

    def _metric(r, label, amount, pct, note, is_key=False, bg=WHITE):
        ws.row_dimensions[r].height = 18
        _cell(ws, r, 1, label, align="left", bg=bg, bold=is_key)
        _cell(ws, r, 2, amount, fmt=GBP, bg=bg, bold=is_key)
        _cell(ws, r, 3, pct,    fmt=PCT, bg=bg, bold=is_key)
        _cell(ws, r, 4, note,   align="left", bg=bg, fg=MUTED_TEXT, italic=True)

    def _spacer(r):
        ws.row_dimensions[r].height = 8

    # ── Headline deduction block ───────────────────────────────────────
    r = 4
    net_row = r
    _metric(r, "Net Revenue (all channels, ex-VAT)",
            ref_mp_net,                                           # cross-ref
            _ratio(2, net_row, 2, net_row),                       # = 100%
            "Shopify Direct + eBay + Amazon + ManoMano + OnBuy + B&Q")
    r += 1
    _spacer(r); r += 1

    fees_row = r
    _metric(r, "Marketplace commissions & fees",
            ref_mp_fees,
            _ratio(2, fees_row, 2, net_row),
            "Referral fees, subscriptions, listing charges — see Marketplace Fees tab")
    r += 1

    comm_row = r
    if summary.get("wayne_commission_overridden"):
        # Audited £ figure — keep as a hard value; suppress the % (em-dash) so
        # funders don't see e.g. 2.9% and ask why it isn't the headline 4%.
        _metric(r, "Commission paid to Wayne Theisinger",
                summary["wayne_commission"], "—",
                summary.get("wayne_commission_note") or "Invoiced Value.")
    else:
        # Auto estimate — show it as a live formula on Net Revenue so the
        # rate is auditable on the face of the sheet.
        _metric(r, "Commission paid to Wayne Theisinger",
                f"={_a1(2, net_row)}*{rate}",
                _ratio(2, comm_row, 2, net_row),
                summary.get("wayne_commission_note")
                    or f"{rate:.0%} of Net Revenue across all channels.")
    r += 1

    ads_row = r
    _metric(r, "Total paid ad spend",
            ref_ads,
            _ratio(2, ads_row, 2, net_row),
            "Google Ads + eBay Promoted Listings")
    r += 1
    _spacer(r); r += 1

    ded_row = r
    _metric(r, "Total deductions (fees + commission + ads)",
            f"={_a1(2, fees_row)}+{_a1(2, comm_row)}+{_a1(2, ads_row)}",
            _ratio(2, ded_row, 2, net_row),
            "All costs combined", is_key=True, bg=LIGHT_GREEN)
    r += 1

    con_row = r
    _metric(r, "Contribution after deductions",
            f"={_a1(2, net_row)}-{_a1(2, ded_row)}",
            _ratio(2, con_row, 2, net_row),
            "Net revenue minus all marketplace costs.", is_key=True, bg=LIGHT_GREEN)
    r += 2

    # Expose headline anchors for the regression test / oracle comparison.
    anchors["summary"] = {
        "net_row": net_row, "fees_row": fees_row, "commission_row": comm_row,
        "ads_row": ads_row, "deductions_row": ded_row, "contribution_row": con_row,
    }

    # ── Commissions & Fees by Channel ──────────────────────────────────
    _section_heading(ws, r, "  Commissions & Fees by Channel", 4)
    r += 1
    _header_row(ws, r, ["Channel", "Net Revenue (£)", "Commissions & Fees (£)", "Fee %"])
    r += 1
    first_ch_row = r
    for ch in data["channels"]:
        bg = PALE_GREEN if r % 2 == 0 else WHITE
        mp = anchors["mp"][ch["name"]]
        _cell(ws, r, 1, ch["name"], align="left", bg=bg, bold=True)
        _cell(ws, r, 2, _xref(SHEET_MARKETPLACE, 4, mp["net_row"]),  fmt=GBP, bg=bg)
        _cell(ws, r, 3, _xref(SHEET_MARKETPLACE, 3, mp["fees_row"]), fmt=GBP, bg=bg)
        _cell(ws, r, 4, _ratio(3, r, 2, r), fmt=PCT, bg=bg)
        r += 1
    last_ch_row = r - 1
    _subtotal_row(ws, r, 4, "TOTAL", {
        2: (_sum(2, first_ch_row, last_ch_row), GBP),
        3: (_sum(3, first_ch_row, last_ch_row), GBP),
        4: (_ratio(3, r, 2, r), PCT),
    })
    r += 2

    # ── Ad Spend by Platform ───────────────────────────────────────────
    # Maps the Summary platform labels onto the Ad Spend tab's platform names
    # so each spend cell can cross-reference the detail tab where possible.
    ad_name_map = {
        "Google Ads":                "Google Ads",
        "eBay Promoted Listings":    "eBay",
        "Amazon Sponsored Products": "Amazon",
    }
    _section_heading(ws, r, "  Ad Spend by Platform", 4)
    r += 1
    _header_row(ws, r, ["Platform", "Spend (£)", "% of Net Revenue", "Note / Coverage"])
    r += 1
    first_ad_row = r
    any_ad_row = False
    for row_data in data.get("ad_spend_platform_summary", []):
        bg = PALE_GREEN if r % 2 == 0 else WHITE
        platform = row_data["platform"]
        ad_tab_name = ad_name_map.get(platform)
        anchor_row  = anchors["ads_platform"].get(ad_tab_name)
        _cell(ws, r, 1, platform, align="left", bg=bg, bold=True)
        if anchor_row is not None:
            _cell(ws, r, 2, _xref(SHEET_ADSPEND, 3, anchor_row), fmt=GBP, bg=bg)
        else:
            # No matching detail-tab cell (e.g. NOT CONNECTED) — fall back to value.
            _cell(ws, r, 2, row_data["spend"], fmt=GBP, bg=bg)
        _cell(ws, r, 3, _ratio(2, r, 2, net_row), fmt=PCT, bg=bg)
        _coverage_cell(ws, r, 4, row_data.get("note"))
        any_ad_row = True
        r += 1
    if any_ad_row:
        _subtotal_row(ws, r, 4, "TOTAL", {
            2: (_sum(2, first_ad_row, r - 1), GBP),
            3: (_ratio(2, r, 2, net_row), PCT),
        })
        r += 2

    if data.get("is_mock"):
        _mock_banner(ws, r, 4)


# ---------------------------------------------------------------------------
# Tab 6 — Raw Data  (pure transaction dump — no derived cells, no formulas)
# ---------------------------------------------------------------------------

def _build_raw(wb: Workbook, data: dict, month_label: str) -> None:
    ws = wb.create_sheet(SHEET_RAW)
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [20, 14, 36, 14, 22, 22])
    _title(ws, f"Raw Data — Transactions  |  {month_label}", 6)
    ws.row_dimensions[2].height = 4

    r = 3

    def write_block(heading: str, col_headers: list[str], rows: list[dict],
                    col_map: list[str], bg_head=MID_GREEN, bg_even=PALE_GREEN) -> int:
        nonlocal r
        _section_heading(ws, r, f"  {heading}", 6, bg=bg_head)
        _header_row(ws, r + 1, col_headers, bg=bg_head)
        r += 2
        for row_data in rows:
            bg = bg_even if r % 2 == 0 else WHITE
            is_ellipsis = str(row_data.get(col_map[0], "")).startswith("…")
            for col_i, key in enumerate(col_map, 1):
                val = row_data.get(key, "")
                c = ws.cell(row=r, column=col_i, value=val)
                c.fill = _fill(bg)
                c.font = Font(size=10, color=MUTED_TEXT if is_ellipsis else DARK_TEXT,
                              italic=is_ellipsis)
                c.alignment = Alignment(
                    horizontal="right" if col_i == 4 and isinstance(val, float) else "left",
                    vertical="center")
                if col_i == 4 and isinstance(val, float):
                    c.number_format = GBP
                c.border = _border_bottom()
            r += 1
        r += 1
        return r

    # eBay
    ebay_raw = data.get("ebay_raw_transactions", [])[:20]
    if ebay_raw:
        write_block(
            "eBay Finance Transactions",
            ["Transaction ID", "Date", "Type", "Amount (£)", "Order ID", "Source"],
            [
                {
                    "id":     t.get("transactionId", ""),
                    "date":   t.get("transactionDate", "")[:10] if t.get("transactionDate") else "",
                    "type":   t.get("transactionType", ""),
                    "amount": float(t.get("amount", {}).get("value", 0)),
                    "order":  t.get("orderId", "—"),
                    "source": "eBay Finances API",
                }
                for t in ebay_raw
            ],
            ["id", "date", "type", "amount", "order", "source"],
        )

    # Amazon — drop zero-amount lines (SP-API returns every fee type even when £0)
    amazon_raw_all = [
        r for r in data.get("amazon_raw_fees", [])
        if abs(float(r.get("amount", 0) or 0)) > 0
    ]
    amazon_raw = amazon_raw_all[:20]
    if amazon_raw:
        write_block(
            "Amazon Settlement Line Items",
            ["Settlement ID", "Posted Date", "Fee Type", "Amount (£)", "Order ID", "Source"],
            amazon_raw,
            ["settlement_id", "posted_at", "fee_type", "amount", "order_id", "source"],
            bg_head=DARK_BLUE, bg_even=PALE_BLUE,
        )

    # BaseLinker — keep only orders where total or commission is non-zero
    # (Amazon fallback orders sometimes have order-level totals stored at
    # product-level instead, leaving the order row empty).
    bl_raw_all = [
        o for o in data.get("baselinker_raw_orders", [])
        if (float(o.get("price_gross", 0) or 0) > 0
            or float(o.get("commission_amount", 0) or 0) > 0)
    ]
    bl_raw = bl_raw_all[:20]
    if bl_raw:
        write_block(
            "BaseLinker Orders (ManoMano / OnBuy)",
            ["Order ID", "Date", "Channel", "Order Total (£)", "Commission (£)", "Source"],
            [
                {
                    "id":         o.get("order_id", ""),
                    "date":       o.get("date_add", ""),
                    "channel":    o.get("order_source", ""),
                    "total":      float(o.get("price_gross", 0) or 0),
                    "commission": float(o.get("commission_amount", 0) or 0),
                    "source":     "BaseLinker",
                }
                for o in bl_raw
            ],
            ["id", "date", "channel", "total", "commission", "source"],
        )

    # Google Ads — drop zero-spend campaigns (the API returns every campaign
    # ever created including paused/archived; only those with spend in the
    # period carry information funders care about).
    google_raw = [
        r_ for r_ in data.get("google_ads_raw", [])
        if float(r_.get("spend_gbp", 0) or 0) > 0
    ]
    if google_raw:
        write_block(
            "Google Ads — Campaign Spend",
            ["Campaign ID", "Type", "Campaign Name", "Spend (£)", "Impressions", "Clicks"],
            [
                {
                    "id":    r_["campaign_id"],
                    "type":  r_["campaign_type"],
                    "name":  r_["campaign_name"],
                    "spend": float(r_["spend_gbp"]),
                    "impr":  r_["impressions"],
                    "click": r_["clicks"],
                }
                for r_ in google_raw
            ],
            ["id", "type", "name", "spend", "impr", "click"],
        )

    if data.get("is_mock"):
        _mock_banner(ws, r, 6)


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def build_workbook(report_data: dict, month_label: str,
                   collect_anchors: dict | None = None) -> Workbook:
    """
    Build the complete 6-tab workbook from assembled report_data.

    Derived cells are written as live Excel formulas; the Summary tab
    cross-references the detail tabs. The detail tabs are built first so their
    cell coordinates are known when the Summary references them, then the
    sheets are reordered so Summary appears first.

    `collect_anchors`: if a dict is provided, it is populated with the key cell
    coordinates the builders register (used by the regression test to locate
    headline cells). Production callers can ignore it.

    report_data keys (all optional — missing data renders as NOT CONNECTED):
        summary              — from transforms.build_summary(); may carry
                               `commission_rate` for the auto-commission formula
        channels             — list of channel dicts with fee_rows
        fba_returns          — payload for the FBA Returns & Removals tab
        fba_returns_note     — coverage note for that tab
        cancellations        — per-channel cancellation aggregates
        cancellation_notes   — coverage notes per channel
        ad_spend_rows        — list of ad spend dicts
        ad_spend_not_connected — list of {platform, note} for missing platforms
        ad_spend_platform_summary — [{platform, spend, note}] for Summary tab
        ad_spend_notes       — {platform: note} for coverage cells
        ebay_raw_transactions
        amazon_raw_fees
        baselinker_raw_orders
        google_ads_raw
        is_mock              — if True, adds amber banner to all tabs
    """
    wb = Workbook()
    wb.remove(wb.active)

    anchors: dict = {"mp": {}, "ads_platform": {}}

    # Detail tabs first (they register the cells the Summary will reference).
    _build_marketplace(wb, report_data, month_label, anchors)
    _build_fba_returns(wb, report_data, month_label, anchors)
    _build_cancellations(wb, report_data, month_label, anchors)
    _build_ad_spend(wb, report_data, month_label, anchors)
    # Summary last — cross-references the detail tabs.
    _build_summary(wb, report_data, month_label, anchors)
    _build_raw(wb, report_data, month_label)

    # Reorder so Summary is the first visible tab.
    order = [SHEET_SUMMARY, SHEET_MARKETPLACE, SHEET_FBA,
             SHEET_CANCEL, SHEET_ADSPEND, SHEET_RAW]
    wb._sheets.sort(key=lambda s: order.index(s.title))

    # Make Excel / Sheets / LibreOffice recalculate every formula on open
    # (openpyxl writes the formula text but does not compute cached values).
    wb.calculation.fullCalcOnLoad = True

    if collect_anchors is not None:
        collect_anchors.update(anchors)

    return wb
