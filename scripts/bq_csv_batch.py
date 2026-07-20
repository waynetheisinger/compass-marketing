"""
Batch CSV generator for the /bq-csv-batch skill.

Pulls MowDirect Shopify products, applies the Shopify->B&Q mapping config
(config/bq_shopify_mapping.json), and produces a Mirakl-importable CSV row
per SKU. One subcategory per batch. The AI work (deciding values, rewriting
prose, fanning out display_attributes JSON) is done by Claude in the skill
loop; this script handles all the deterministic plumbing.

Subcommands:
    list-templates
        Scan bq/templates/*.xlsx, emit JSON of templates + their subcategories.
        The skill's Q1 picker is built from this output.

    init-batch --category "..." --skus "S1 S2 ..." [--batch-name X]
        Validate category exists in a template. Resolve the template file +
        subcategory column. Create batch state, empty CSV (headers only),
        empty log. Returns batch_id and paths.

    check-metafields --skus "S1 S2 ..."
        Pull metafields for each SKU. Cross-check against the mapping config.
        Emit list of metafields (namespace.key) seen on these SKUs that are
        not in config/bq_shopify_mapping.json. The skill surfaces these so
        Claude can add a mapping entry inline (or rely on judgement during
        the per-SKU pass, which reads the full metafield set regardless).

    gather --batch-state PATH --sku SKU
        For one SKU, pull Shopify data + deref metaobjects, subset the mapping
        config to relevant entries, extract template column metadata for the
        chosen subcategory (required/recommended/NA, value-lists, units).
        Emit one big JSON file at bq/csv-output/<batch>_<sku>_context.json
        that Claude reads to do the AI step.

    write-row --batch-state PATH --sku SKU --row-json PATH
        Take Claude's filled-row JSON, run deterministic validator (value-list
        membership, numeric parse, name char-cap), demote failures to UNKNOWN,
        append the row to the CSV, append per-SKU section to log.md, update
        batch state stats.

    finalize --batch-state PATH
        Print summary; show output paths.

Batch state shape (bq/csv-output/.state_<batch>.json):
    {
      "batch_id":              "<batch_name>_<YYYYMMDD_HHMM>",
      "batch_name":            "<batch_name>",
      "created_at":            "2026-05-20T15:30:00Z",
      "subcategory":           "Pressure Washers (Cordless)",
      "subcategory_full":      "Marketplace Categories/Cleaning & Waste/...",
      "template_file":         "bq/templates/pressure_washers.xlsx",
      "subcategory_col_index": 5,
      "csv_path":              "bq/csv-output/<batch_id>.csv",
      "log_path":              "bq/csv-output/<batch_id>.log.md",
      "skus":                  ["S1", "S2", ...],
      "processed_skus":        [],
      "stats":                 {"cells_filled": 0, "cells_unknown": 0, "validator_failures": 0}
    }
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import warnings
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

# Suppress openpyxl's default-style warning that Mirakl's templates trigger
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
if _THIS_DIR not in sys.path:
    sys.path.insert(0, _THIS_DIR)

import marketplace_dims  # noqa: E402  (needs _THIS_DIR on sys.path)

_REPO_ROOT = os.path.dirname(_THIS_DIR)
_TEMPLATES_DIR = os.path.join(_REPO_ROOT, "bq/templates")
_OUTPUT_DIR = os.path.join(_REPO_ROOT, "bq/csv-output")
_UPSERT_DIR = os.path.join(_REPO_ROOT, "bq/csv-upsert")
_MAPPING_PATH = os.path.join(_REPO_ROOT, "config", "bq_shopify_mapping.json")
_PIM_MAP_PATH = os.path.join(_REPO_ROOT, "config", "bq_subcategory_pim_map.json")
_VLIST_CODES_PATH = os.path.join(_REPO_ROOT, "config", "bq_value_list_codes.json")

# Banned characters per BQ_QUIRKS — full em-dash, en-dash, multiplication sign,
# degree sign, smart quotes. The 130-char name cap is also a BQ_QUIRKS rule.
_BANNED_CHARS = "—–×°‘’“”"
_NAME_MAX_LEN = 130


# ---------------------------------------------------------------------------
# Time + state helpers
# ---------------------------------------------------------------------------

def _utcnow_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M")


def _slugify(s: str) -> str:
    s = re.sub(r"[^a-zA-Z0-9_-]+", "-", s.strip().lower())
    return re.sub(r"-+", "-", s).strip("-") or "batch"


def _load_state(path: str) -> dict[str, Any]:
    with open(path) as fh:
        return json.load(fh)


def _save_state(path: str, state: dict[str, Any]) -> None:
    with open(path, "w") as fh:
        json.dump(state, fh, indent=2, ensure_ascii=False)
        fh.write("\n")


def _load_vlist_codes() -> dict[str, Any]:
    """Curated label->code map keyed by CSV column api_code.

    Shape:
      {
        "global":  {<api_code>: {<label>: <code>}},
        "by_pim":  {"PIM_xxxx": {<api_code>: {<label>: <code>}}},
      }
    Keys starting with '_' (e.g. _comment, _lookup_order) are ignored.
    """
    if not os.path.isfile(_VLIST_CODES_PATH):
        return {"global": {}, "by_pim": {}}
    with open(_VLIST_CODES_PATH) as fh:
        raw = json.load(fh)
    return {
        "global": raw.get("global", {}),
        "by_pim": raw.get("by_pim", {}),
    }


def _translate_label_to_code(
    api_code: str, label: str, pim: str | None, vlist_codes: dict[str, Any]
) -> str | None:
    """Look up the Mirakl code for (api_code, label).

    Lookup order: by_pim[pim][api_code] -> global[api_code] -> None.
    Returns None if no translation is configured (caller leaves the label
    as-is).
    """
    if pim:
        pim_section = vlist_codes.get("by_pim", {}).get(pim, {})
        if api_code in pim_section and label in pim_section[api_code]:
            return pim_section[api_code][label]
    global_section = vlist_codes.get("global", {})
    if api_code in global_section and label in global_section[api_code]:
        return global_section[api_code][label]
    return None


def _load_pim_map() -> dict[str, str]:
    """Subcategory breadcrumb -> Kingfisher PIM hierarchy code.

    Mirakl's product import expects the PIM code (e.g. 'PIM_13946') in
    column A, not the breadcrumb. The breadcrumb produces error
    1001/1004 ('category unknown / could not be identified'). Keys
    starting with '_' are treated as comments.
    """
    if not os.path.isfile(_PIM_MAP_PATH):
        return {}
    with open(_PIM_MAP_PATH) as fh:
        raw = json.load(fh)
    return {k: v for k, v in raw.items() if not k.startswith("_")}


def _persist_pim_map_entry(breadcrumb: str, pim_code: str) -> None:
    """Append (breadcrumb -> pim_code) to config/bq_subcategory_pim_map.json,
    preserving comment keys and key order (alphabetical for non-comments)."""
    raw: dict[str, str] = {}
    if os.path.isfile(_PIM_MAP_PATH):
        with open(_PIM_MAP_PATH) as fh:
            raw = json.load(fh)
    comments = {k: v for k, v in raw.items() if k.startswith("_")}
    entries = {k: v for k, v in raw.items() if not k.startswith("_")}
    entries[breadcrumb] = pim_code
    merged: dict[str, str] = {}
    merged.update(comments)
    for k in sorted(entries):
        merged[k] = entries[k]
    with open(_PIM_MAP_PATH, "w") as fh:
        json.dump(merged, fh, indent=2, ensure_ascii=False)
        fh.write("\n")


def _resolve_pim_from_mirakl(breadcrumb: str) -> str | None:
    """Fetch /hierarchies from Mirakl Kingfisher and resolve breadcrumb ->
    PIM_* code. Returns None if the API is unreachable, credentials are
    missing, or the breadcrumb doesn't match any hierarchy entry.

    Walks each entry's parent chain to reconstruct its full breadcrumb and
    matches against the target. Only entries with code prefix 'PIM_' are
    considered (Mirakl's product import only accepts these as category
    values; other prefixes like 1P_ROOT are structural).
    """
    try:
        from scripts.mirakl_client import MiraklClient
    except Exception:
        return None
    try:
        client = MiraklClient("KINGFISHER")
    except Exception:
        return None
    try:
        resp = client.get("/hierarchies")
    except Exception:
        return None
    items = resp.get("hierarchies", []) if isinstance(resp, dict) else []
    by_code = {h["code"]: h for h in items if "code" in h}

    def crumb(code: str) -> str:
        parts: list[str] = []
        cur = by_code.get(code)
        while cur:
            parts.append(cur.get("label", ""))
            parent = cur.get("parent_code")
            cur = by_code.get(parent) if parent else None
        return "/".join(reversed(parts))

    for h in items:
        code = h.get("code", "")
        if not code.startswith("PIM_"):
            continue
        if crumb(code) == breadcrumb:
            return code
    return None


def _load_mapping() -> dict[str, Any]:
    if not os.path.exists(_MAPPING_PATH):
        raise SystemExit(
            "config/bq_shopify_mapping.json does not exist. "
            "Restore it from git (git checkout config/bq_shopify_mapping.json)."
        )
    with open(_MAPPING_PATH) as fh:
        return json.load(fh)


# ---------------------------------------------------------------------------
# Template reading
# ---------------------------------------------------------------------------

def _open_template(path: str):
    import openpyxl

    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def _list_templates() -> list[dict[str, Any]]:
    """Scan bq/templates/*.xlsx, return list of {file, subcategories,
    subcategory_indices}. `subcategories` is the user-facing list (no
    Nones, stable order). `subcategory_indices` maps each subcategory
    string to its actual column index in the Columns sheet — some
    templates contain a `None` separator cell between metadata and
    subcategory columns (e.g. Lawnmowers_without_Batteries), so the
    on-disk index is not always `4 + position_in_subcategories`.
    """
    out: list[dict[str, Any]] = []
    if not os.path.isdir(_TEMPLATES_DIR):
        return out
    for name in sorted(os.listdir(_TEMPLATES_DIR)):
        if not name.endswith(".xlsx"):
            continue
        path = os.path.join(_TEMPLATES_DIR, name)
        wb = _open_template(path)
        try:
            ws = wb["Columns"]
            header = next(ws.iter_rows(values_only=True))
            # Columns sheet row 0: ('Code', 'Label', 'Description', 'Value example', [maybe None separator,] '<subcat1>', '<subcat2>', ...)
            indices = {h: i for i, h in enumerate(header[4:], start=4) if h}
            subcategories = list(indices.keys())
            out.append({
                "file": os.path.relpath(path, _REPO_ROOT),
                "subcategories": subcategories,
                "subcategory_indices": indices,
            })
        finally:
            wb.close()
    return out


def _resolve_subcategory(category_name: str) -> tuple[str, int, str]:
    """Find which template contains category_name. Returns (template_path,
    column_index_in_Columns_sheet, exact_subcategory_string).

    Matching is loose: case-insensitive substring + exact match wins ties.
    """
    matches: list[tuple[str, int, str, bool]] = []  # (path, col_idx, name, exact)
    for t in _list_templates():
        path = os.path.join(_REPO_ROOT, t["file"])
        for sub in t["subcategories"]:
            exact = sub == category_name
            partial = category_name.lower() in sub.lower()
            if exact or partial:
                col_idx = t["subcategory_indices"][sub]
                matches.append((path, col_idx, sub, exact))
    if not matches:
        raise SystemExit(
            f"No subcategory matches '{category_name}' in any bq/templates/*.xlsx. "
            f"Available subcategories: run `list-templates` to see."
        )
    # Prefer exact match if any
    exact = [m for m in matches if m[3]]
    if exact:
        path, idx, name, _ = exact[0]
        return path, idx, name
    if len(matches) > 1:
        names = [m[2] for m in matches]
        raise SystemExit(
            f"Ambiguous category '{category_name}' matched: {names}. "
            f"Be more specific."
        )
    path, idx, name, _ = matches[0]
    return path, idx, name


def _extract_column_meta(template_path: str, subcat_col_index: int) -> list[dict[str, Any]]:
    """Read the Columns sheet of the template and build per-column metadata
    for the chosen subcategory.

    Returns a list of dicts in the order they appear in the Data sheet:
        {
          "api_code": "shop_sku",
          "portal_label": "Shop SKU",
          "description": "...",
          "required": "REQUIRED" | "RECOMMENDED" | "NA" | None,
          "value_list": ["Li-ion", ...] or None,
          "unit": "kg" or None,
          "numeric_only": bool
        }
    """
    wb = _open_template(template_path)
    try:
        # Pull Columns sheet rows
        ws_cols = wb["Columns"]
        rows = list(ws_cols.iter_rows(values_only=True))
        # rows[0] is header, rows[1:] are entries
        col_entries: dict[str, dict[str, Any]] = {}
        for row in rows[1:]:
            api_code = row[0]
            if not api_code:
                continue
            col_entries[api_code] = {
                "api_code": api_code,
                "portal_label": row[1],
                "description": row[2] or "",
                "required": row[subcat_col_index] if subcat_col_index < len(row) else None,
            }

        # Pull Data sheet to know the order + canonical API code list
        ws_data = wb["Data"]
        data_rows = list(ws_data.iter_rows(values_only=True))
        # Row 0 = portal labels; Row 1 = API codes
        api_codes_in_order = [c for c in data_rows[1] if c]

        # Pull ReferenceData for value-lists
        ws_ref = wb["ReferenceData"]
        ref_rows = list(ws_ref.iter_rows(values_only=True))
        # ref_rows[0] = (category, ...attribute_api_codes...)
        ref_header = ref_rows[0]
        value_lists: dict[str, list[str]] = {}
        for col_idx, api_code in enumerate(ref_header):
            if col_idx == 0 or not api_code:
                continue
            values: list[str] = []
            for r in ref_rows[1:]:
                if col_idx < len(r) and r[col_idx]:
                    values.append(str(r[col_idx]))
            if values:
                value_lists[api_code] = values

        # Assemble in Data sheet order
        out: list[dict[str, Any]] = []
        for api_code in api_codes_in_order:
            entry = col_entries.get(api_code, {
                "api_code": api_code,
                "portal_label": api_code,
                "description": "",
                "required": None,
            })
            desc = entry["description"]
            # Heuristic: unit detection from description like "Hose length (m)"
            # or "Please provide numeric values only, kg will be applied automatically"
            unit_match = re.search(r"\(([a-zA-Z°²³µ/]+)\)\s*$", entry["portal_label"] or "")
            unit = unit_match.group(1) if unit_match else None
            if not unit:
                unit_match2 = re.search(r"\b([a-zA-Z°]+)\s+will be applied automatically", desc)
                unit = unit_match2.group(1) if unit_match2 else None
            numeric_only = "numeric values only" in desc.lower()
            entry.update({
                "value_list": value_lists.get(api_code),
                "unit": unit,
                "numeric_only": numeric_only,
            })
            out.append(entry)
        return out
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Upsert input — existing B&Q catalogue values for safety
# ---------------------------------------------------------------------------

def _load_upsert_index() -> dict[str, dict[str, dict[str, str]]]:
    """Scan bq/csv-upsert/*.{csv,xlsx} for downloaded B&Q product exports.

    Returns:
        {
          "by_ean":      {"5065...": {api_code: value, ...}, ...},
          "by_shop_sku": {"SBSCBC":  {api_code: value, ...}, ...},
          "sources":     {(key_kind, key): "filename.xlsx", ...},
        }

    Empty cells are dropped (so a missing column does not look like an
    intentional blank). Headers are expected to be B&Q api_codes; xlsx files
    with a 'Data' sheet (Mirakl template/export shape) use row 1 (api codes),
    skipping row 0 (portal labels). Otherwise row 0 is treated as the header.
    """
    by_ean: dict[str, dict[str, str]] = {}
    by_shop_sku: dict[str, dict[str, str]] = {}
    sources: dict[tuple[str, str], str] = {}

    if not os.path.isdir(_UPSERT_DIR):
        return {"by_ean": by_ean, "by_shop_sku": by_shop_sku, "sources": sources}

    for entry in sorted(os.listdir(_UPSERT_DIR)):
        if entry.startswith(".") or entry.lower().endswith((".md", ".txt", ".log")):
            continue
        path = os.path.join(_UPSERT_DIR, entry)
        if not os.path.isfile(path):
            continue
        ext = entry.lower().rsplit(".", 1)[-1]
        try:
            rows = _read_upsert_file(path, ext)
        except Exception as e:
            print(f"[warn] Could not parse upsert file {entry}: {e}", file=sys.stderr)
            continue
        for header, data_rows in rows:
            ean_idx = _header_index(header, "ean")
            sku_idx = _header_index(header, "shop_sku")
            for r in data_rows:
                cells: dict[str, str] = {}
                for i, h in enumerate(header):
                    if not h:
                        continue
                    if i >= len(r):
                        continue
                    v = r[i]
                    if v is None or (isinstance(v, str) and not v.strip()):
                        continue
                    cells[str(h).strip()] = str(v).strip() if isinstance(v, str) else str(v)
                if not cells:
                    continue
                if ean_idx is not None and ean_idx < len(r) and r[ean_idx]:
                    key = str(r[ean_idx]).strip()
                    by_ean[key] = {**by_ean.get(key, {}), **cells}
                    sources[("EAN", key)] = entry
                if sku_idx is not None and sku_idx < len(r) and r[sku_idx]:
                    key = str(r[sku_idx]).strip()
                    by_shop_sku[key] = {**by_shop_sku.get(key, {}), **cells}
                    sources[("SHOP_SKU", key)] = entry

    return {"by_ean": by_ean, "by_shop_sku": by_shop_sku, "sources": sources}


def _read_upsert_file(path: str, ext: str) -> list[tuple[list[str], list[list[Any]]]]:
    """Return [(header_row, data_rows), ...] for an upsert file.

    Most files yield one (header, rows) tuple. XLSX with a 'Data' sheet uses
    row index 1 as the header (api codes), matching Mirakl template shape.
    """
    if ext == "csv":
        with open(path, newline="", encoding="utf-8-sig") as fh:
            reader = csv.reader(fh)
            all_rows = list(reader)
        if not all_rows:
            return []
        # Mirakl exports use a TWO-row header: row 0 = portal labels
        # (e.g. "Shop SKU", "EAN"), row 1 = api codes (e.g. "shop_sku",
        # "ean"). Detect this by looking for api-code style identifiers
        # on row 1; if present, use row 1 as the header and skip row 0.
        header_idx = _detect_header_row(all_rows)
        return [(all_rows[header_idx], all_rows[header_idx + 1:])]
    if ext in ("xlsx", "xlsm"):
        wb = _open_template(path)
        try:
            sheets = wb.sheetnames
            if "Data" in sheets:
                ws = wb["Data"]
                rows = [list(r) for r in ws.iter_rows(values_only=True)]
                if len(rows) >= 2:
                    return [(rows[1], rows[2:])]
                return []
            ws = wb[sheets[0]]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            if not rows:
                return []
            header_idx = _detect_header_row(rows)
            return [(rows[header_idx], rows[header_idx + 1:])]
        finally:
            wb.close()
    raise ValueError(f"unsupported extension: {ext}")


def _detect_header_row(rows: list[list[Any]]) -> int:
    """Return index of the api-code header row.

    Mirakl exports have row 0 = portal labels ("Shop SKU", "EAN") and
    row 1 = api codes ("shop_sku", "ean"). If row 1 contains api-code
    identifiers, use it; otherwise row 0 is the header.
    """
    api_code_markers = {"shop_sku", "ean", "category"}
    if len(rows) >= 2:
        row1 = {str(c).strip().lower() for c in rows[1] if c}
        if api_code_markers & row1:
            return 1
    return 0


def _header_index(header: list[Any], key: str) -> int | None:
    """Case-insensitive search for `key` in a header row; returns column index."""
    k = key.strip().lower()
    for i, h in enumerate(header):
        if h and str(h).strip().lower() == k:
            return i
    return None


def _lookup_existing(
    upsert_index: dict[str, dict[str, dict[str, str]]],
    sku: str,
    ean: str | None,
) -> tuple[dict[str, str], str | None, str | None]:
    """Return (row_dict, match_kind, source_file).

    EAN match first, then shop_sku. Returns ({}, None, None) on miss.
    """
    if ean:
        row = upsert_index["by_ean"].get(str(ean).strip())
        if row:
            src = upsert_index["sources"].get(("EAN", str(ean).strip()))
            return row, "EAN", src
    if sku:
        row = upsert_index["by_shop_sku"].get(str(sku).strip())
        if row:
            src = upsert_index["sources"].get(("SHOP_SKU", str(sku).strip()))
            return row, "SHOP_SKU", src
    return {}, None, None


# ---------------------------------------------------------------------------
# Shopify pull
# ---------------------------------------------------------------------------

def _shopify_pull_full(sku: str) -> dict[str, Any] | None:
    """Pull a single product (by variant SKU) with everything we'll need."""
    from shopify_client import ShopifyClient

    query = """
    query($q: String!) {
      productVariants(first: 5, query: $q) {
        edges {
          node {
            sku
            barcode
            price
            inventoryQuantity
            inventoryItem { measurement { weight { value unit } } }
            metafields(first: 100) {
              edges { node { namespace key type value } }
            }
            product {
              id
              title
              descriptionHtml
              vendor
              productType
              tags
              status
              featuredImage { url altText }
              images(first: 20) { edges { node { url altText } } }
              category { id name fullName }
              metafields(first: 100) {
                edges { node { namespace key type value } }
              }
            }
          }
        }
      }
    }
    """
    with ShopifyClient() as client:
        res = client.execute(query, variables={"q": f"sku:{sku}"})
    for e in res["productVariants"]["edges"]:
        if e["node"]["sku"] == sku:
            return e["node"]
    return None


def _shopify_deref_metaobjects(gids: list[str]) -> dict[str, str]:
    """Resolve metaobject gids to display labels."""
    from shopify_client import ShopifyClient

    if not gids:
        return {}
    query = """
    query($ids: [ID!]!) {
      nodes(ids: $ids) {
        ... on Metaobject {
          id
          displayName
          handle
          fields { key value }
        }
      }
    }
    """
    out: dict[str, str] = {}
    with ShopifyClient() as client:
        for i in range(0, len(gids), 50):
            chunk = gids[i : i + 50]
            res = client.execute(query, variables={"ids": chunk})
            for node in res["nodes"]:
                if not node:
                    continue
                gid = node["id"]
                display = node.get("displayName") or node.get("handle") or gid
                for f in node.get("fields") or []:
                    if f["key"] in ("label", "name") and f.get("value"):
                        display = f["value"]
                        break
                out[gid] = display
    return out


def _flatten_metafields(variant: dict[str, Any], gid_lookup: dict[str, str]) -> dict[str, dict[str, Any]]:
    """Build {namespace.key: {type, value, dereffed_value, owner_type}} for all
    metafields on the variant + product."""
    out: dict[str, dict[str, Any]] = {}
    for owner_label, mf_block in (
        ("PRODUCTVARIANT", variant.get("metafields", {}).get("edges", [])),
        ("PRODUCT", variant.get("product", {}).get("metafields", {}).get("edges", [])),
    ):
        for edge in mf_block:
            mf = edge["node"]
            key = f"{mf['namespace']}.{mf['key']}"
            raw = mf["value"]
            mtype = mf["type"]
            derefed: Any = raw
            if mtype == "list.metaobject_reference" and raw:
                try:
                    gids = json.loads(raw)
                    derefed = [gid_lookup.get(g, g) for g in gids]
                except (json.JSONDecodeError, TypeError):
                    pass
            elif mtype == "metaobject_reference" and raw:
                derefed = gid_lookup.get(raw, raw)
            elif mtype == "json" and raw:
                try:
                    derefed = json.loads(raw)
                except (json.JSONDecodeError, TypeError):
                    pass
            out[key] = {
                "owner_type": owner_label,
                "type": mtype,
                "raw_value": raw,
                "value": derefed,
            }
    return out


# ---------------------------------------------------------------------------
# Subcommands
# ---------------------------------------------------------------------------

def _cmd_list_templates() -> int:
    templates = _list_templates()
    all_subs: list[dict[str, Any]] = []
    for t in templates:
        for sub in t["subcategories"]:
            all_subs.append({"subcategory": sub, "template": t["file"]})
    print(json.dumps({"templates": templates, "all_subcategories": all_subs}, indent=2, ensure_ascii=False))
    return 0


def _cmd_init_batch(category: str, skus: list[str], batch_name: str | None) -> int:
    template_path, col_idx, subcat_exact = _resolve_subcategory(category)
    stamp = _stamp()
    name = batch_name or _slugify(subcat_exact)
    batch_id = f"{name}_{stamp}"

    state_path = os.path.join(_OUTPUT_DIR, f".state_{batch_id}.json")
    csv_path = os.path.join(_OUTPUT_DIR, f"{batch_id}.csv")
    log_path = os.path.join(_OUTPUT_DIR, f"{batch_id}.log.md")

    columns = _extract_column_meta(template_path, col_idx)
    api_codes = [c["api_code"] for c in columns]

    state = {
        "batch_id": batch_id,
        "batch_name": name,
        "created_at": _utcnow_iso(),
        "subcategory": subcat_exact,
        "subcategory_full": subcat_exact,  # full path lives in the template
        "template_file": os.path.relpath(template_path, _REPO_ROOT),
        "subcategory_col_index": col_idx,
        "csv_path": os.path.relpath(csv_path, _REPO_ROOT),
        "log_path": os.path.relpath(log_path, _REPO_ROOT),
        "skus": skus,
        "processed_skus": [],
        "stats": {"cells_filled": 0, "cells_unknown": 0, "validator_failures": 0},
        "column_count": len(columns),
        "required_columns": [c["api_code"] for c in columns if c["required"] == "REQUIRED"],
    }
    os.makedirs(_OUTPUT_DIR, exist_ok=True)
    _save_state(state_path, state)

    # Write CSV header — single row of api codes. The two-row header
    # (portal labels + api codes) was tested 2026-05-20 and made things
    # worse: Mirakl treats row 1 as the header (using portal labels) and
    # row 2 (api codes) as a data row, producing an extra error.
    with open(csv_path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(api_codes)

    # Write log header
    with open(log_path, "w") as fh:
        fh.write(f"# B&Q CSV Batch — {name}, {stamp}\n")
        fh.write(f"Template: {os.path.relpath(template_path, _REPO_ROOT)}\n")
        fh.write(f"Subcategory: {subcat_exact}\n")
        fh.write(f"SKUs: {len(skus)}\n")
        fh.write(f"Columns: {len(columns)} ({len(state['required_columns'])} REQUIRED)\n\n")

    print(json.dumps({
        "ok": True,
        "batch_id": batch_id,
        "state_path": os.path.relpath(state_path, _REPO_ROOT),
        "csv_path": state["csv_path"],
        "log_path": state["log_path"],
        "template": state["template_file"],
        "subcategory": subcat_exact,
        "column_count": len(columns),
        "required_count": len(state["required_columns"]),
    }, indent=2))
    return 0


def _cmd_check_metafields(skus: list[str]) -> int:
    """Pull metafields for each SKU; report any namespace.key not in mapping."""
    mapping = _load_mapping()
    known_product = set(mapping["shopify_metafields"]["PRODUCT"].keys())
    known_variant = set(mapping["shopify_metafields"]["PRODUCTVARIANT"].keys())

    summary: dict[str, Any] = {
        "skus_processed": 0,
        "skus_not_found": [],
        "unmapped_metafields": {"PRODUCT": [], "PRODUCTVARIANT": []},
    }
    seen_unmapped: dict[str, set[str]] = {"PRODUCT": set(), "PRODUCTVARIANT": set()}

    from shopify_client import ShopifyClient
    query = """
    query($q: String!) {
      productVariants(first: 5, query: $q) {
        edges {
          node {
            sku
            metafields(first: 100) { edges { node { namespace key } } }
            product { metafields(first: 100) { edges { node { namespace key } } } }
          }
        }
      }
    }
    """
    with ShopifyClient() as client:
        for sku in skus:
            res = client.execute(query, variables={"q": f"sku:{sku}"})
            found = False
            for e in res["productVariants"]["edges"]:
                if e["node"]["sku"] != sku:
                    continue
                found = True
                summary["skus_processed"] += 1
                for mf_e in e["node"]["product"]["metafields"]["edges"]:
                    mf = mf_e["node"]
                    key = f"{mf['namespace']}.{mf['key']}"
                    if key not in known_product:
                        seen_unmapped["PRODUCT"].add(key)
                for mf_e in e["node"]["metafields"]["edges"]:
                    mf = mf_e["node"]
                    key = f"{mf['namespace']}.{mf['key']}"
                    if key not in known_variant:
                        seen_unmapped["PRODUCTVARIANT"].add(key)
            if not found:
                summary["skus_not_found"].append(sku)

    summary["unmapped_metafields"]["PRODUCT"] = sorted(seen_unmapped["PRODUCT"])
    summary["unmapped_metafields"]["PRODUCTVARIANT"] = sorted(seen_unmapped["PRODUCTVARIANT"])
    summary["has_unmapped_metafields"] = bool(seen_unmapped["PRODUCT"] or seen_unmapped["PRODUCTVARIANT"])
    print(json.dumps(summary, indent=2))
    return 0


def _cmd_gather(state_path: str, sku: str) -> int:
    state = _load_state(state_path)
    mapping = _load_mapping()

    variant = _shopify_pull_full(sku)
    if variant is None:
        print(json.dumps({"error": f"SKU {sku} not found on Shopify"}), file=sys.stderr)
        return 2

    # Collect metaobject GIDs across this product, deref in one round-trip
    all_mf_edges: list[dict[str, Any]] = []
    all_mf_edges.extend(e["node"] for e in variant.get("metafields", {}).get("edges", []))
    all_mf_edges.extend(e["node"] for e in variant.get("product", {}).get("metafields", {}).get("edges", []))
    gids_to_deref: list[str] = []
    for mf in all_mf_edges:
        if mf["type"] == "metaobject_reference" and mf["value"]:
            gids_to_deref.append(mf["value"])
        elif mf["type"] == "list.metaobject_reference" and mf["value"]:
            try:
                gids_to_deref.extend(json.loads(mf["value"]))
            except (json.JSONDecodeError, TypeError):
                pass
    gid_lookup = _shopify_deref_metaobjects(gids_to_deref) if gids_to_deref else {}
    metafields = _flatten_metafields(variant, gid_lookup)

    # Template metadata
    template_path = os.path.join(_REPO_ROOT, state["template_file"])
    columns = _extract_column_meta(template_path, state["subcategory_col_index"])
    # Filter out NA columns from the AI's view — they don't apply to this subcategory
    active_columns = [c for c in columns if c["required"] != "NA"]

    # Subset mapping config to entries that target columns in this subcategory
    active_api_codes = {c["api_code"] for c in active_columns}
    relevant_mapping: dict[str, dict[str, Any]] = {}
    for section_name in ("shopify_core_fields", "shopify_metafields"):
        for owner in ("PRODUCT", "PRODUCTVARIANT"):
            for key, entry in mapping[section_name][owner].items():
                if entry.get("skip_reason"):
                    continue
                if not entry.get("bq_targets"):
                    continue
                if not set(entry["bq_targets"]) & active_api_codes:
                    continue
                relevant_mapping[f"{section_name}.{owner}.{key}"] = {
                    "section": section_name,
                    "owner_type": owner,
                    "shopify_key": key,
                    "type": entry.get("type"),
                    "bq_targets": entry["bq_targets"],
                    "hint": entry.get("hint"),
                    "note": entry.get("note"),
                }

    # Shopify data — strip giant noise (e.g. judgeme HTML), keep useful slices
    shopify_summary = {
        "sku": variant["sku"],
        "barcode": variant.get("barcode"),
        "price": variant.get("price"),
        "inventoryQuantity": variant.get("inventoryQuantity"),
        "weight": (variant.get("inventoryItem") or {}).get("measurement", {}).get("weight"),
        "product": {
            "title": variant["product"]["title"],
            "descriptionHtml": variant["product"]["descriptionHtml"],
            "vendor": variant["product"]["vendor"],
            "productType": variant["product"]["productType"],
            "tags": variant["product"]["tags"],
            "status": variant["product"]["status"],
            "category": variant["product"].get("category"),
            "featuredImage": variant["product"].get("featuredImage"),
            "images": [e["node"] for e in variant["product"]["images"]["edges"]],
        },
        "metafields": metafields,
    }

    # Look up existing B&Q values for this SKU (safety net for upserts).
    # Match on EAN first, fall back to shop_sku. Subset to columns active in
    # this subcategory so the AI doesn't see noise from other categories.
    upsert_index = _load_upsert_index()
    existing_row, match_kind, match_source = _lookup_existing(
        upsert_index, sku, variant.get("barcode")
    )
    existing_values: dict[str, str] = {}
    if existing_row:
        for code in active_api_codes:
            if code in existing_row:
                existing_values[code] = existing_row[code]

    context = {
        "sku": sku,
        "batch_id": state["batch_id"],
        "subcategory": state["subcategory"],
        "shopify": shopify_summary,
        "mapping_config_relevant": relevant_mapping,
        "bq_columns": active_columns,
        "existing_b_and_q_values": existing_values,
        "existing_b_and_q_match": (
            {"match_kind": match_kind, "source_file": match_source}
            if existing_values
            else None
        ),
        "context_files_to_read": [
            ".claude/skills/enrich-bq/BRAND_VOICE.md",
            ".claude/skills/enrich-bq/BQ_QUIRKS.md",
        ],
        "row_output_schema": {
            "filled":  "{<api_code>: <value as string>, ...}  — only fill cells you're confident about",
            "unknown": "[{column: <api_code>, reason: <short reason>}, ...]  — every cell you can't fill confidently",
        },
        "validator_will_check": {
            "value_list_fields": "value must appear in the column's value_list array (case sensitive)",
            "numeric_fields":    "value must parse as a number (just digits + optional decimal point)",
            "name_field":        f"<= {_NAME_MAX_LEN} chars, no banned chars: em-dash, en-dash, ×, °, smart quotes",
        },
        "upsert_safety_note": (
            "write-row auto-restores any active column that you mark UNKNOWN or "
            "omit if existing_b_and_q_values has a value for it. You do NOT need "
            "to copy these into 'filled' manually — the script handles it and "
            "tags them [RESTORED] in the log. Override only when you have a "
            "better value (i.e. fill the cell yourself)."
        ) if existing_values else None,
        "warnings": [],
    }

    if (variant.get("inventoryQuantity") or 0) <= 0:
        context["warnings"].append(f"SKU {sku} has inventoryQuantity={variant.get('inventoryQuantity')} (out of stock)")
    if variant["product"]["productType"]:
        # Loose check: does productType share a token with subcategory?
        pt_tokens = set(re.findall(r"\w+", (variant["product"]["productType"] or "").lower()))
        sub_tokens = set(re.findall(r"\w+", state["subcategory"].lower()))
        if not pt_tokens & sub_tokens:
            context["warnings"].append(
                f"SKU {sku} productType='{variant['product']['productType']}' shares no tokens with "
                f"subcategory='{state['subcategory']}'. Confirm this SKU belongs in this batch."
            )

    context_path = os.path.join(_OUTPUT_DIR, f"{state['batch_id']}_{sku}_context.json")
    with open(context_path, "w") as fh:
        json.dump(context, fh, indent=2, ensure_ascii=False)

    print(json.dumps({
        "ok": True,
        "sku": sku,
        "context_path": os.path.relpath(context_path, _REPO_ROOT),
        "shopify_data_present": True,
        "active_columns_count": len(active_columns),
        "required_count": sum(1 for c in active_columns if c["required"] == "REQUIRED"),
        "recommended_count": sum(1 for c in active_columns if c["required"] == "RECOMMENDED"),
        "relevant_mapping_count": len(relevant_mapping),
        "existing_values_count": len(existing_values),
        "existing_values_match_kind": match_kind,
        "existing_values_source": match_source,
        "warnings": context["warnings"],
    }, indent=2))
    return 0


# ---------------------------------------------------------------------------
# Validator
# ---------------------------------------------------------------------------

def _validate_cell(api_code: str, value: str, column_meta: dict[str, Any]) -> tuple[bool, str | None]:
    """Returns (passes, reason_if_fail). Value is a string."""
    if value is None or value == "":
        return True, None  # Empty is OK; only filled values get validated
    s = str(value)
    if api_code == "name":
        if len(s) > _NAME_MAX_LEN:
            return False, f"'name' is {len(s)} chars; max {_NAME_MAX_LEN}"
        bad = [c for c in s if c in _BANNED_CHARS]
        if bad:
            return False, f"'name' contains banned chars per BQ_QUIRKS: {sorted(set(bad))}"
    if column_meta.get("value_list"):
        if s not in column_meta["value_list"]:
            # Helpful nearby suggestions
            lower = s.lower()
            near = [v for v in column_meta["value_list"] if lower in v.lower() or v.lower() in lower][:3]
            hint = f"; nearest in value list: {near}" if near else ""
            return False, f"'{api_code}': value '{s}' not in value list{hint}"
    if column_meta.get("numeric_only"):
        try:
            float(s)
        except ValueError:
            return False, f"'{api_code}' expects numeric only; got '{s}'"
    return True, None


def _cmd_write_row(state_path: str, sku: str, row_json_path: str) -> int:
    state = _load_state(state_path)
    if sku in state["processed_skus"]:
        print(json.dumps({"error": f"SKU {sku} already processed in this batch"}), file=sys.stderr)
        return 2

    with open(row_json_path) as fh:
        row = json.load(fh)
    filled: dict[str, str] = row.get("filled") or {}
    unknown: list[dict[str, str]] = row.get("unknown") or []

    # Load template column metadata to validate against
    template_path = os.path.join(_REPO_ROOT, state["template_file"])
    columns = _extract_column_meta(template_path, state["subcategory_col_index"])
    col_meta_by_code = {c["api_code"]: c for c in columns}
    api_codes_in_order = [c["api_code"] for c in columns]
    required_codes = {c["api_code"] for c in columns if c["required"] == "REQUIRED"}
    recommended_codes = {c["api_code"] for c in columns if c["required"] == "RECOMMENDED"}

    # Validate. Failures get demoted to unknown.
    validator_failures: list[dict[str, str]] = []
    cleaned_filled: dict[str, str] = {}
    for api_code, value in filled.items():
        if api_code not in col_meta_by_code:
            validator_failures.append({"column": api_code, "reason": f"column not in this subcategory's template"})
            continue
        if col_meta_by_code[api_code]["required"] == "NA":
            validator_failures.append({"column": api_code, "reason": f"column is NA for this subcategory"})
            continue
        passes, reason = _validate_cell(api_code, value, col_meta_by_code[api_code])
        if not passes:
            validator_failures.append({"column": api_code, "reason": reason, "rejected_value": str(value)})
            unknown.append({"column": api_code, "reason": f"VALIDATOR: {reason}"})
        else:
            cleaned_filled[api_code] = str(value) if value is not None else ""

    # Set column A (category) FIRST so the upsert restore loop sees it as
    # already filled. Mirakl wants the Kingfisher PIM code (e.g. PIM_13946),
    # not the breadcrumb path. PIM map: config/bq_subcategory_pim_map.json.
    # On cache miss, fetch /hierarchies from Mirakl and persist the entry so
    # subsequent runs are offline-cached.
    pim_map = _load_pim_map()
    pim_code = pim_map.get(state["subcategory"])
    category_warning: str | None = None
    pim_autofetched = False
    if not pim_code:
        pim_code = _resolve_pim_from_mirakl(state["subcategory"])
        if pim_code:
            _persist_pim_map_entry(state["subcategory"], pim_code)
            pim_autofetched = True
    if pim_code:
        cleaned_filled["category"] = pim_code
    else:
        wb = _open_template(template_path)
        try:
            header = next(wb["Columns"].iter_rows(values_only=True))
            category_path = header[state["subcategory_col_index"]] or state["subcategory"]
        finally:
            wb.close()
        cleaned_filled["category"] = category_path
        category_warning = (
            f"No PIM code for subcategory '{state['subcategory']}' in "
            f"config/bq_subcategory_pim_map.json, and Mirakl /hierarchies "
            f"auto-fetch failed (credentials missing, API unreachable, or "
            f"breadcrumb not found). Falling back to breadcrumb "
            f"'{category_path}' (Mirakl will likely reject with error "
            f"1001/1004). Add the PIM code to the config manually and re-run."
        )

    # Upsert safety net: any active column that the AI did NOT fill (left
    # UNKNOWN or omitted) gets back-filled from the existing B&Q export, if
    # one matches this SKU. Restored values are re-validated; failures are
    # dropped (the cell stays blank) and logged so we never re-assert junk.
    upsert_index = _load_upsert_index()
    sku_barcode = None
    # Pull EAN from the row first (most reliable — AI just put it there), else
    # the state file's batch metadata won't have it; that's fine, shop_sku is
    # the fallback.
    if "ean" in cleaned_filled and cleaned_filled["ean"]:
        sku_barcode = cleaned_filled["ean"]
    existing_row, match_kind, match_source = _lookup_existing(
        upsert_index, sku, sku_barcode
    )
    restored: list[dict[str, str]] = []
    restore_failures: list[dict[str, str]] = []
    if existing_row:
        unknown_codes = {u["column"] for u in unknown}
        for col in columns:
            code = col["api_code"]
            if col["required"] == "NA":
                continue
            if code in cleaned_filled and cleaned_filled[code]:
                continue  # AI filled it — keep AI's value
            if code not in existing_row:
                continue
            val = existing_row[code]
            passes, reason = _validate_cell(code, val, col)
            if not passes:
                restore_failures.append({"column": code, "reason": reason, "rejected_value": val})
                continue
            cleaned_filled[code] = val
            restored.append({"column": code, "value": val})
            # If it was in the unknown list, remove it — it's no longer unknown
            if code in unknown_codes:
                unknown = [u for u in unknown if u["column"] != code]

    # Translate value-list labels into Mirakl codes BEFORE writing the row.
    # Mirakl rejects labels with error 2006; codes are required for
    # value-list-backed attributes (e.g. "Spectrum" -> "4592", "240V" -> "16",
    # "5 years" -> "81"). Lookup table: config/bq_value_list_codes.json.
    # PIM-specific codes override globals; if no mapping exists for a cell,
    # the label is left as-is and the validator's earlier check ensures it
    # at least appears in the template's value_list (Mirakl will still
    # reject — but the demotion-to-UNKNOWN path would have already caught
    # totally invalid values).
    vlist_codes = _load_vlist_codes()
    label_to_code_translations: list[dict[str, str]] = []
    untranslated_value_list_cells: list[dict[str, str]] = []
    for code_col, label_value in list(cleaned_filled.items()):
        if not label_value:
            continue
        column_meta = col_meta_by_code.get(code_col)
        if not column_meta or not column_meta.get("value_list"):
            continue
        translated = _translate_label_to_code(code_col, label_value, pim_code, vlist_codes)
        if translated is not None and translated != label_value:
            cleaned_filled[code_col] = translated
            label_to_code_translations.append({
                "column": code_col, "label": label_value, "code": translated,
            })
        elif translated is None:
            # No mapping configured for this column. The label might be
            # accepted (some value-list fields take literal strings, e.g.
            # reach_verified="Yes") or rejected by Mirakl. Flag it so Wayne
            # can extend the map if needed.
            untranslated_value_list_cells.append({"column": code_col, "label": label_value})

    # Build row in API code order
    out_row = [cleaned_filled.get(code, "") for code in api_codes_in_order]
    csv_path = os.path.join(_REPO_ROOT, state["csv_path"])
    with open(csv_path, "a", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(out_row)

    # Physical product/shipping dimensions blank → surface as gaps regardless of
    # the template's REQUIRED marker (they must never regress silently). Runs
    # after the upsert restore + value-list translation so restored dims count.
    dimension_gaps = marketplace_dims.missing_dimension_gaps(columns, cleaned_filled)
    for gap in dimension_gaps:
        if not any(u["column"] == gap["column"] for u in unknown):
            unknown.append(gap)
    if dimension_gaps:
        print(f"[DIMENSION] {sku}: {len(dimension_gaps)} physical dimension "
              f"cell(s) blank — {', '.join(g['column'] for g in dimension_gaps)}",
              file=sys.stderr)

    # Build log section
    log_path = os.path.join(_REPO_ROOT, state["log_path"])
    cells_filled = sum(1 for v in cleaned_filled.values() if v)
    cells_unknown = len(unknown)
    cells_restored = len(restored)
    with open(log_path, "a") as fh:
        fh.write(f"\n## {sku}\n")
        fh.write(
            f"Filled: {cells_filled} cells   "
            f"Restored: {cells_restored} cells   "
            f"Unknown: {cells_unknown} cells   "
            f"Validator failures: {len(validator_failures)}"
        )
        if match_kind:
            fh.write(f"   Upsert match: {match_kind} from {match_source}")
        fh.write("\n\n")
        if category_warning:
            fh.write(f"### [CATEGORY WARNING]\n- {category_warning}\n\n")
        if label_to_code_translations:
            fh.write(f"### Label -> code translations ({len(label_to_code_translations)})\n")
            for t in label_to_code_translations:
                fh.write(f"- [CODE] {t['column']} — {t['label']!r} -> {t['code']!r}\n")
            fh.write("\n")
        if untranslated_value_list_cells:
            fh.write(f"### Untranslated value-list cells ({len(untranslated_value_list_cells)})\n")
            for u in untranslated_value_list_cells:
                fh.write(f"- [NO_CODE_MAP] {u['column']} — {u['label']!r} sent as-is (Mirakl may reject)\n")
            fh.write("\n")
        if restored:
            fh.write(f"### Restored from upsert ({match_kind} {match_source})\n")
            for r in restored:
                col = r["column"]
                tag = "[REQUIRED]" if col in required_codes else ("[RECOMMENDED]" if col in recommended_codes else "[OPTIONAL]")
                val_preview = r["value"] if len(r["value"]) <= 80 else r["value"][:77] + "..."
                fh.write(f"- [RESTORED] {tag} {col} — {val_preview}\n")
            fh.write("\n")
        if restore_failures:
            fh.write(f"### Restore-attempt validator failures ({match_kind} {match_source})\n")
            for rf in restore_failures:
                fh.write(f"- [RESTORE_REJECTED] {rf['column']} — {rf['reason']} (was: {rf.get('rejected_value','')})\n")
            fh.write("\n")
        if unknown:
            fh.write("### Gaps (cells still blank)\n")
            for u in unknown:
                col = u["column"]
                tag = "[REQUIRED]" if col in required_codes else ("[RECOMMENDED]" if col in recommended_codes else "[OPTIONAL]")
                if str(u.get("reason", "")).startswith("VALIDATOR:"):
                    tag = "[VALIDATOR]"
                if u.get("dimension"):
                    tag = "[DIMENSION]"
                fh.write(f"- {tag} {col} — {u.get('reason') or 'no reason given'}\n")

    # Update state
    state["processed_skus"].append(sku)
    state["stats"]["cells_filled"] += cells_filled
    state["stats"]["cells_unknown"] += cells_unknown
    state["stats"]["validator_failures"] += len(validator_failures)
    state["stats"].setdefault("cells_restored", 0)
    state["stats"]["cells_restored"] += cells_restored
    state["stats"].setdefault("dimensions_blank", 0)
    state["stats"]["dimensions_blank"] += len(dimension_gaps)
    _save_state(state_path, state)

    print(json.dumps({
        "ok": True,
        "sku": sku,
        "cells_filled": cells_filled,
        "cells_restored": cells_restored,
        "cells_unknown": cells_unknown,
        "dimensions_blank": [g["column"] for g in dimension_gaps],
        "validator_failures": validator_failures,
        "upsert_match": {"kind": match_kind, "source": match_source} if match_kind else None,
        "restored_columns": [r["column"] for r in restored],
        "restore_validator_failures": restore_failures,
        "category_value": cleaned_filled["category"],
        "category_warning": category_warning,
        "pim_autofetched": pim_autofetched,
        "label_to_code_translations": label_to_code_translations,
        "untranslated_value_list_cells": untranslated_value_list_cells,
        "progress": f"{len(state['processed_skus'])}/{len(state['skus'])}",
    }, indent=2))
    return 0


def _cmd_finalize(state_path: str) -> int:
    state = _load_state(state_path)
    missing = [s for s in state["skus"] if s not in state["processed_skus"]]
    print(json.dumps({
        "ok": True,
        "batch_id": state["batch_id"],
        "csv_path": state["csv_path"],
        "log_path": state["log_path"],
        "skus_processed": len(state["processed_skus"]),
        "skus_missing": missing,
        "stats": state["stats"],
    }, indent=2))
    return 0


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(prog="bq_csv_batch")
    sub = parser.add_subparsers(dest="cmd", required=True)

    sub.add_parser("list-templates", help="List templates + subcategories")

    init = sub.add_parser("init-batch", help="Initialize batch state, CSV header, log header")
    init.add_argument("--category", required=True, help="Exact or unique-substring B&Q subcategory")
    init.add_argument("--skus", required=True, help="Space-separated SKU list")
    init.add_argument("--batch-name", default=None)

    chk = sub.add_parser("check-metafields", help="Find metafields on these SKUs not in config")
    chk.add_argument("--skus", required=True, help="Space-separated SKU list")

    gat = sub.add_parser("gather", help="Gather context for one SKU")
    gat.add_argument("--batch-state", required=True)
    gat.add_argument("--sku", required=True)

    wrt = sub.add_parser("write-row", help="Validate + append one row to CSV/log")
    wrt.add_argument("--batch-state", required=True)
    wrt.add_argument("--sku", required=True)
    wrt.add_argument("--row-json", required=True)

    fin = sub.add_parser("finalize", help="Print batch summary")
    fin.add_argument("--batch-state", required=True)

    args = parser.parse_args(argv)
    if args.cmd == "list-templates":
        return _cmd_list_templates()
    if args.cmd == "init-batch":
        skus = args.skus.split()
        if not skus:
            parser.error("--skus is empty")
        return _cmd_init_batch(args.category, skus, args.batch_name)
    if args.cmd == "check-metafields":
        skus = args.skus.split()
        return _cmd_check_metafields(skus)
    if args.cmd == "gather":
        return _cmd_gather(args.batch_state, args.sku)
    if args.cmd == "write-row":
        return _cmd_write_row(args.batch_state, args.sku, args.row_json)
    if args.cmd == "finalize":
        return _cmd_finalize(args.batch_state)
    parser.error(f"Unknown command: {args.cmd}")
    return 2


if __name__ == "__main__":
    sys.exit(main())
