"""
Dump the structure + every populated cell from Chris's
"Marketing Spend v Return.xlsx" so we can see what fields he expects us
to fill in before Monday.
"""
from openpyxl import load_workbook

PATH = "/Users/waynetheisinger/compass/marketingPlan/Marketing Spend v Return.xlsx"

wb = load_workbook(PATH, data_only=False)
print(f"Sheets: {wb.sheetnames}\n")

for name in wb.sheetnames:
    ws = wb[name]
    print("=" * 80)
    print(f"SHEET: {name!r}   dims={ws.dimensions}   "
          f"max_row={ws.max_row}  max_col={ws.max_column}")
    print("=" * 80)

    # Merged cells
    if ws.merged_cells.ranges:
        print(f"  Merged ranges ({len(ws.merged_cells.ranges)}):")
        for r in ws.merged_cells.ranges:
            print(f"    {r}")
        print()

    # Every populated cell — show coord, formula (if any), value
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is None or (isinstance(v, str) and not v.strip()):
                continue
            extra = ""
            if isinstance(v, str) and v.startswith("="):
                extra = "  [FORMULA]"
            print(f"  {cell.coordinate:<6}  {v!r}{extra}")
    print()

# Also load with values resolved (no formulas) so we see what Excel last computed.
print("=" * 80)
print("VALUE-RESOLVED VIEW (data_only=True)")
print("=" * 80)
wb2 = load_workbook(PATH, data_only=True)
for name in wb2.sheetnames:
    ws = wb2[name]
    has_any = any(c.value is not None for row in ws.iter_rows() for c in row)
    if not has_any:
        continue
    print(f"\n-- {name!r} --")
    for row in ws.iter_rows():
        line = []
        for cell in row:
            v = cell.value
            if v is None:
                line.append("")
            else:
                line.append(str(v))
        if any(x.strip() for x in line):
            print("  " + " | ".join(line))
