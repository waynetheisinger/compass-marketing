"""
Generate a mock monthly spend report spreadsheet for funder review.
Uses realistic figures derived from the MowDirect 2026 marketing plan.
Run: python scripts/mock_report.py
Output: reports/marketing_spend_mock_2026-03.xlsx
"""
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Mock data — March 2026
# Figures derived from marketing plan rates and seasonal model
# ---------------------------------------------------------------------------

MONTH_LABEL = "March 2026"
GROSS_TURNOVER = 435_635.00

# Each channel has:
#   fees      — marketing/commission costs (cost of selling on the platform)
#   fba_costs — Amazon FBA operational costs (cost of sales, NOT marketing)
#   ad_spend  — paid advertising spend with impressions/clicks for the Ad Spend tab

CHANNELS = [
    {
        "name": "Shopify Direct",
        "source": "Shopify GraphQL (mock)",
        "net": 174_254.00,
        "fees": [
            ("Payment processing (Shopify Payments 1.5%)", 2_614.00),
        ],
        "fba_costs": [],
        "ad_spend": [],
    },
    {
        "name": "eBay",
        "source": "eBay Finances API (mock)",
        "net": 108_909.00,
        "fees": [
            ("Referral / final value fees (11%)",  11_980.00),
            ("Promoted Listings spend",              2_178.00),
            ("Shipping label costs",                   891.00),
            ("Store subscription",                      49.00),
        ],
        "fba_costs": [],
        "ad_spend": [
            ("eBay Promoted Listings", 2_178.00, 412_000, 5_340),
        ],
    },
    {
        "name": "Amazon",
        "source": "SP-API Settlement (mock)",
        "net": 87_127.00,
        # Referral fee is a commission (marketing cost of selling on Amazon)
        "fees": [
            ("Referral fees / commission (12%)",  10_455.00),
        ],
        # FBA operational costs — separate from marketing budget
        "fba_costs": [
            ("FBA fulfilment fees (pick, pack & ship)",     8_713.00),
            ("FBA storage fees (standard, March)",          1_218.00),
            ("FBA inbound shipping (stock to warehouse)",   1_480.00),
            ("FBA prep & labelling",                          348.00),
        ],
        "ad_spend": [
            ("Amazon Sponsored Products", 3_050.00, 189_000, 3_210),
        ],
    },
    {
        "name": "ManoMano",
        "source": "BaseLinker (mock)",
        "net": 34_851.00,
        "fees": [
            ("Commission (18%)", 6_273.00),
        ],
        "fba_costs": [],
        "ad_spend": [],
    },
    {
        "name": "OnBuy",
        "source": "BaseLinker (mock)",
        "net": 17_425.00,
        "fees": [
            ("Commission (7%)", 1_220.00),
        ],
        "fba_costs": [],
        "ad_spend": [],
    },
    {
        "name": "B&Q (Mirakl)",
        "source": "Mirakl API (mock)",
        "net": 13_069.00,
        "fees": [
            ("Commission (13%)",            1_699.00),
            ("Listing / platform charges",    120.00),
        ],
        "fba_costs": [],
        "ad_spend": [],
    },
]

GOOGLE_ADS = [
    ("Performance Max — All Products",   10_673.00, 1_240_000, 18_650),
    ("Shopping — Spectrum Own-Label",     4_574.00,   380_000,  6_920),
]

# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

DARK_GREEN  = "1B5E20"
MID_GREEN   = "2E7D32"
LIGHT_GREEN = "C8E6C9"
PALE_GREEN  = "E8F5E9"
DARK_BLUE   = "0D47A1"
MID_BLUE    = "1565C0"
LIGHT_BLUE  = "BBDEFB"
PALE_BLUE   = "E3F2FD"
WHITE       = "FFFFFF"
AMBER       = "FFF8E1"
DARK_TEXT   = "212121"
MUTED_TEXT  = "757575"

GBP = '£#,##0.00'
PCT = '0.0%'
INT = '#,##0'


def fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)


def bold_font(size=11, colour=DARK_TEXT):
    return Font(bold=True, size=size, color=colour)


def normal_font(size=11, colour=DARK_TEXT, italic=False):
    return Font(size=size, color=colour, italic=italic)


def thin_border_bottom():
    return Border(bottom=Side(style="thin", color="E0E0E0"))


def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def header_row(ws, row, values, bg=DARK_GREEN, fg=WHITE):
    ws.row_dimensions[row].height = 22
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.font = Font(bold=True, size=10, color=fg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def data_cell(ws, row, col, value, fmt=None, bold_=False, bg=WHITE,
              align="right", fg=DARK_TEXT, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    if fmt:
        c.number_format = fmt
    c.font = Font(bold=bold_, size=10, color=fg, italic=italic)
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = thin_border_bottom()
    return c


def section_heading(ws, row, text, ncols, bg=MID_GREEN):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.row_dimensions[row].height = 20
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill(bg)
    c.font = Font(bold=True, size=10, color=WHITE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def mock_banner(ws, row, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.row_dimensions[row].height = 22
    c = ws.cell(row=row, column=1,
                value="⚠  MOCK DATA — illustrative figures based on 2026 plan projections. "
                      "Replace with live API data before sharing with funders.")
    c.fill = fill(AMBER)
    c.font = Font(bold=True, size=9, color="E65100")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)


def title_row(ws, text, ncols, bg=DARK_GREEN):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.row_dimensions[1].height = 34
    c = ws["A1"]
    c.value = text
    c.fill = fill(bg)
    c.font = Font(bold=True, size=13, color=WHITE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def subtotal_row(ws, row, ncols, label, col_values, bg=LIGHT_GREEN,
                 fg=DARK_TEXT, accent_bg=None):
    """Write a subtotal row. col_values = {col_index: (value, fmt)}"""
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill(bg)
    data_cell(ws, row, 1, label, align="right", bg=bg, bold_=True, fg=fg)
    for col, (val, fmt) in col_values.items():
        data_cell(ws, row, col, val, fmt=fmt, bg=bg, bold_=True, fg=fg)


def grand_total_row(ws, row, ncols, label, col_values, bg=DARK_GREEN):
    ws.row_dimensions[row].height = 18
    for col in range(1, ncols + 1):
        ws.cell(row=row, column=col).fill = fill(bg)
        ws.cell(row=row, column=col).font = Font(bold=True, size=10, color=WHITE)
    data_cell(ws, row, 1, label, align="left", bg=bg, bold_=True, fg=WHITE)
    for col, (val, fmt) in col_values.items():
        data_cell(ws, row, col, val, fmt=fmt, bg=bg, bold_=True, fg=WHITE)
        ws.cell(row=row, column=col).font = Font(bold=True, size=10, color=WHITE)


# ---------------------------------------------------------------------------
# Derived totals
# ---------------------------------------------------------------------------

WAYNE_COMMISSION_RATE = 0.04  # 4% of total net revenue, replaces FBA cost-of-sales


def totals():
    """
    Mock totals. Channel values in CHANNELS are illustrative round numbers —
    treated here as the **net** (ex-VAT) revenue base for the funder
    report. No 1.20 back-out applied to keep the mock readable.
    """
    all_fees      = sum(f for ch in CHANNELS for _, f in ch["fees"])
    all_google    = sum(s for _, s, _, _ in GOOGLE_ADS)
    all_mkt_ads   = sum(s for ch in CHANNELS for _, s, _, _ in ch["ad_spend"])
    all_ads       = all_google + all_mkt_ads
    all_net       = sum(ch["net"] for ch in CHANNELS)
    commission    = round(all_net * WAYNE_COMMISSION_RATE, 2)
    return {
        "net":        all_net,
        "fees":       all_fees,
        "commission": commission,
        "ads":        all_ads,
        "combined":   all_fees + commission + all_ads,
    }


# ---------------------------------------------------------------------------
# Tab 1 — Summary
# ---------------------------------------------------------------------------

def build_summary(wb):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [36, 18, 16, 40])
    title_row(ws, f"MowDirect — Monthly Cost Report  |  {MONTH_LABEL}", 4)

    ws.row_dimensions[2].height = 4  # spacer

    t = totals()

    # ── Key metrics table ────────────────────────────────────────────────────
    header_row(ws, 3, ["Metric", "Amount (£)", "% of Net Revenue", "Notes"])

    metrics = [
        ("Net Revenue (all channels, ex-VAT)", t["net"], t["net"] / t["net"],
         "Shopify + eBay + Amazon + ManoMano + OnBuy + B&Q"),
        None,
        ("Marketplace commissions & fees",   t["fees"],     t["fees"]  / t["net"],
         "Referral fees, subscriptions, listing charges — see Marketplace Fees tab"),
        ("Commission paid to Wayne Theisinger", t["commission"], t["commission"] / t["net"],
         "4% of Net Revenue across all channels."),
        ("Total paid ad spend",              t["ads"],      t["ads"]   / t["net"],
         "Google Ads + eBay Promoted Listings + Amazon Sponsored Products"),
        None,
        ("Total deductions (fees + commission + ads)", t["combined"], t["combined"] / t["net"],
         "All costs combined"),
        ("Contribution after deductions",    t["net"] - t["combined"],
         (t["net"] - t["combined"]) / t["net"],
         "Net revenue minus all marketplace costs."),
    ]

    r = 4
    shading = [PALE_GREEN, WHITE]
    shade_i = 0
    for m in metrics:
        if m is None:
            ws.row_dimensions[r].height = 8
            r += 1
            continue
        label, amount, pct, note = m
        is_key = label.startswith("Total deductions") or label.startswith("Contribution")
        bg = LIGHT_GREEN if is_key else shading[shade_i % 2]
        shade_i += 1
        ws.row_dimensions[r].height = 18
        data_cell(ws, r, 1, label,  align="left", bg=bg, bold_=is_key)
        data_cell(ws, r, 2, amount, fmt=GBP,       bg=bg, bold_=is_key)
        data_cell(ws, r, 3, pct,    fmt=PCT,        bg=bg, bold_=is_key)
        data_cell(ws, r, 4, note,   align="left",  bg=bg, fg=MUTED_TEXT, italic=True)
        r += 1

    r += 1

    # ── Two-column breakdown: commissions vs FBA ─────────────────────────────
    # Left: commission by channel  |  Right: FBA by cost type
    section_heading(ws, r, "  Commissions & Fees by Channel", 4)
    r += 1
    header_row(ws, r, ["Channel", "Net Revenue (£)", "Commissions & Fees (£)", "Fee %"])
    r += 1
    for ch in CHANNELS:
        ch_fees = sum(f for _, f in ch["fees"])
        bg = PALE_GREEN if r % 2 == 0 else WHITE
        data_cell(ws, r, 1, ch["name"],  align="left", bg=bg, bold_=True)
        data_cell(ws, r, 2, ch["net"], fmt=GBP, bg=bg)
        data_cell(ws, r, 3, ch_fees,     fmt=GBP, bg=bg)
        data_cell(ws, r, 4, ch_fees / ch["net"], fmt=PCT, bg=bg)
        r += 1
    subtotal_row(ws, r, 4, "TOTAL",
                 {2: (t["net"], GBP), 3: (t["fees"], GBP),
                  4: (t["fees"] / t["net"], PCT)})
    r += 2

    section_heading(ws, r, "  Ad Spend by Platform", 4)
    r += 1
    header_row(ws, r, ["Platform", "Spend (£)", "% of Net Revenue", "Notes"])
    r += 1
    all_google = sum(s for _, s, _, _ in GOOGLE_ADS)
    ad_rows = [
        ("Google Ads (PMax + Shopping)", all_google,
         "Campaign detail in Ad Spend tab"),
        ("eBay Promoted Listings",
         sum(s for _, s, _, _ in CHANNELS[1]["ad_spend"]),
         "Charged as AD_FEE via eBay Finances API"),
        ("Amazon Sponsored Products",
         sum(s for _, s, _, _ in CHANNELS[2]["ad_spend"]),
         "Via Amazon Advertising API"),
    ]
    for i, (platform, spend, note) in enumerate(ad_rows):
        bg = PALE_GREEN if r % 2 == 0 else WHITE
        data_cell(ws, r, 1, platform, align="left", bg=bg, bold_=True)
        data_cell(ws, r, 2, spend, fmt=GBP, bg=bg)
        data_cell(ws, r, 3, spend / t["net"], fmt=PCT, bg=bg)
        data_cell(ws, r, 4, note, align="left", bg=bg, fg=MUTED_TEXT, italic=True)
        r += 1
    subtotal_row(ws, r, 4, "TOTAL",
                 {2: (t["ads"], GBP), 3: (t["ads"] / t["net"], PCT)})
    r += 2

    mock_banner(ws, r, 4)


# ---------------------------------------------------------------------------
# Tab 2 — Marketplace Fees & Commissions  (marketing budget view)
# ---------------------------------------------------------------------------

def build_marketplace(wb):
    ws = wb.create_sheet("Marketplace Fees & Commissions")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [24, 36, 14, 14, 14, 26, 16])
    title_row(ws, f"Marketplace Commissions & Fees  |  {MONTH_LABEL}", 7)
    ws.row_dimensions[2].height = 4

    header_row(ws, 3,
               ["Channel", "Fee Type", "Amount (£)", "Net Channel Revenue (£)",
                "Fee as % of Net Revenue", "Data Source", "Coverage"])

    t = totals()
    r = 4
    for ch in CHANNELS:
        ch_fees = sum(f for _, f in ch["fees"])
        for j, (fee_type, amount) in enumerate(ch["fees"]):
            bg = PALE_GREEN if r % 2 == 0 else WHITE
            data_cell(ws, r, 1, ch["name"] if j == 0 else "",
                      align="left", bg=bg, bold_=(j == 0))
            data_cell(ws, r, 2, fee_type, align="left", bg=bg)
            data_cell(ws, r, 3, amount, fmt=GBP, bg=bg)
            data_cell(ws, r, 4, ch["net"] if j == 0 else None,
                      fmt=GBP, bg=bg)
            data_cell(ws, r, 5, amount / ch["net"], fmt=PCT, bg=bg)
            data_cell(ws, r, 6, ch["source"] if j == 0 else "",
                      align="left", bg=bg, fg=MUTED_TEXT, italic=True)
            data_cell(ws, r, 7, "MOCK DATA", align="center",
                      bg=AMBER, fg="E65100", bold_=True)
            r += 1
        # Channel subtotal
        subtotal_row(ws, r, 7, f"{ch['name']} subtotal",
                     {3: (ch_fees, GBP), 4: (ch["net"], GBP),
                      5: (ch_fees / ch["net"], PCT)})
        r += 1

    grand_total_row(ws, r, 7, "GRAND TOTAL — MARKETPLACE FEES",
                   {3: (t["fees"], GBP), 4: (t["net"], GBP),
                    5: (t["fees"] / t["net"], PCT)})
    r += 2
    mock_banner(ws, r, 7)


# ---------------------------------------------------------------------------
# Tab — Ad Spend
# (FBA Cost of Sales tab removed — see git history for the prior layout.)
# ---------------------------------------------------------------------------

def build_ad_spend(wb):
    ws = wb.create_sheet("Ad Spend")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [24, 38, 14, 14, 10, 10, 20])
    title_row(ws, f"Paid Ad Spend  |  {MONTH_LABEL}", 7)
    ws.row_dimensions[2].height = 4

    header_row(ws, 3,
               ["Platform", "Campaign", "Spend (£)", "Impressions",
                "Clicks", "CTR", "Coverage"])

    t = totals()
    r = 4

    # Google Ads
    for j, (camp, spend, impr, clicks) in enumerate(GOOGLE_ADS):
        bg = PALE_GREEN if r % 2 == 0 else WHITE
        data_cell(ws, r, 1, "Google Ads" if j == 0 else "",
                  align="left", bg=bg, bold_=(j == 0))
        data_cell(ws, r, 2, camp, align="left", bg=bg)
        data_cell(ws, r, 3, spend,  fmt=GBP, bg=bg)
        data_cell(ws, r, 4, impr,   fmt=INT, bg=bg)
        data_cell(ws, r, 5, clicks, fmt=INT, bg=bg)
        data_cell(ws, r, 6, clicks / impr, fmt=PCT, bg=bg)
        data_cell(ws, r, 7, "MOCK DATA", align="center",
                  bg=AMBER, fg="E65100", bold_=True)
        r += 1

    g_total = sum(s for _, s, _, _ in GOOGLE_ADS)
    g_impr  = sum(i for _, _, i, _ in GOOGLE_ADS)
    g_click = sum(c for _, _, _, c in GOOGLE_ADS)
    subtotal_row(ws, r, 7, "Google Ads subtotal",
                 {3: (g_total, GBP), 4: (g_impr, INT),
                  5: (g_click, INT), 6: (g_click / g_impr, PCT)})
    r += 1

    # Per-channel ad spend
    for ch in CHANNELS:
        if not ch["ad_spend"]:
            continue
        for j, (camp, spend, impr, clicks) in enumerate(ch["ad_spend"]):
            bg = PALE_GREEN if r % 2 == 0 else WHITE
            data_cell(ws, r, 1, ch["name"], align="left", bg=bg, bold_=True)
            data_cell(ws, r, 2, camp, align="left", bg=bg)
            data_cell(ws, r, 3, spend,  fmt=GBP, bg=bg)
            data_cell(ws, r, 4, impr,   fmt=INT, bg=bg)
            data_cell(ws, r, 5, clicks, fmt=INT, bg=bg)
            data_cell(ws, r, 6, clicks / impr, fmt=PCT, bg=bg)
            data_cell(ws, r, 7, "MOCK DATA", align="center",
                      bg=AMBER, fg="E65100", bold_=True)
            r += 1

    r += 1
    grand_total_row(ws, r, 7, "TOTAL AD SPEND",
                   {3: (t["ads"], GBP)})
    r += 2
    mock_banner(ws, r, 7)


# ---------------------------------------------------------------------------
# Tab 5 — Raw Data (sample rows)
# ---------------------------------------------------------------------------

def build_raw(wb):
    ws = wb.create_sheet("Raw Data (samples)")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [18, 14, 36, 14, 20, 20])
    title_row(ws, f"Raw Data — Sample Transactions  |  {MONTH_LABEL}", 6)
    ws.row_dimensions[2].height = 4

    def write_sample_block(ws, start_row, heading, col_headers, rows, bg_even, bg_head=MID_GREEN):
        section_heading(ws, start_row, f"  {heading}", 6, bg=bg_head)
        header_row(ws, start_row + 1, col_headers, bg=bg_head)
        r = start_row + 2
        for row_data in rows:
            is_ellipsis = row_data[0] == "…"
            bg = bg_even if r % 2 == 0 else WHITE
            for col, val in enumerate(row_data, 1):
                c = ws.cell(row=r, column=col, value=val)
                c.fill = fill(PALE_BLUE if bg_head == DARK_BLUE else bg)
                c.font = normal_font(colour=MUTED_TEXT, italic=True) if is_ellipsis \
                    else normal_font()
                c.alignment = Alignment(
                    horizontal="right" if col == 4 else "left",
                    vertical="center")
                if col == 4 and isinstance(val, float):
                    c.number_format = GBP
            r += 1
        return r + 1

    r = 3
    r = write_sample_block(ws, r, "eBay Finance Transactions",
        ["Transaction ID", "Date", "Type", "Amount (£)", "Order ID", "Source"],
        [
            ("EBAY-FEE-001", "2026-03-01", "FINAL_VALUE_FEE",  -45.23, "28-12345-67890", "eBay Finances API"),
            ("EBAY-FEE-002", "2026-03-02", "FINAL_VALUE_FEE",  -67.81, "28-12346-67891", "eBay Finances API"),
            ("EBAY-AD-001",  "2026-03-03", "AD_FEE",           -12.40, "28-12345-67890", "eBay Finances API"),
            ("EBAY-SHP-001", "2026-03-04", "SHIPPING_LABEL",    -8.95, "28-12347-67892", "eBay Finances API"),
            ("EBAY-SUB-001", "2026-03-01", "SUBSCRIPTION_FEE", -49.00, "—",              "eBay Finances API"),
            ("…", "…", "… (all transactions for month)", "…", "…", ""),
        ],
        PALE_GREEN)

    r = write_sample_block(ws, r, "Amazon Settlement Line Items",
        ["Settlement ID", "Posted Date", "Fee Type", "Amount (£)", "Order ID", "Source"],
        [
            ("S-12345678", "2026-03-01", "Referral Fee",           -52.14, "123-4567890-1234567", "SP-API Settlement"),
            ("S-12345678", "2026-03-01", "FBA Fulfillment Fee",    -11.35, "123-4567890-1234567", "SP-API Settlement"),
            ("S-12345678", "2026-03-02", "Referral Fee",           -38.70, "123-4567891-1234568", "SP-API Settlement"),
            ("S-12345679", "2026-03-15", "FBA Storage Fee",       -407.00, "—",                   "SP-API Settlement"),
            ("S-12345679", "2026-03-15", "FBA Inbound Transport",  -93.50, "—",                   "SP-API Settlement"),
            ("…", "…", "… (all line items for month)", "…", "…", ""),
        ],
        PALE_BLUE, bg_head=DARK_BLUE)

    r = write_sample_block(ws, r, "Google Ads — Campaign Spend",
        ["Campaign ID", "Type", "Campaign Name", "Spend (£)", "Impressions", "Clicks"],
        [
            ("12345001", "PERFORMANCE_MAX", "PMax — All Products",       10_673.00, 1_240_000, 18_650),
            ("12345002", "SHOPPING",        "Shopping — Spectrum Range",  4_574.00,   380_000,  6_920),
        ],
        PALE_GREEN)

    mock_banner(ws, r, 6)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    os.makedirs("reports", exist_ok=True)
    out = "reports/marketing_spend_mock_2026-03.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    build_summary(wb)
    build_marketplace(wb)
    build_ad_spend(wb)
    build_raw(wb)

    wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    main()
