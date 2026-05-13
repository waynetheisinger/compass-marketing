"""
Populate Chris's 'Marketing Spend v Return.xlsx' with MowDirect's API-driven
figures, in place. Backs up the original first.

Fills:
  Row  6  Google Ads (Mow)     — B (March) · C (April) · D (May 1–11)
  Row 22  Mow Website          — B (March) · C (April) · D (May 1–11)
  Row 26  Amazon FBA Sales     — B (March) · C (April) · D (May 1–11)

Note: the May column is month-to-date through 11 May 2026 to match the
'Compass Website' partial-May figure already in the workbook (£77,841 for
~9 days). Add a note in Wayne's hand-off if Chris wanted a different cut-off.
"""
from __future__ import annotations
import shutil
import time
from datetime import datetime, date, timezone
from pathlib import Path

from openpyxl import load_workbook

from scripts.google_ads_client import GoogleAdsClient
from scripts.shopify_client import ShopifyClient
from scripts.amazon_client import AmazonClient

WB_PATH = Path("/Users/waynetheisinger/compass/marketingPlan/Marketing Spend v Return.xlsx")
BACKUP  = WB_PATH.with_suffix(".backup-2026-05-11.xlsx")

# Month windows (May is partial — month-to-date through 11 May)
WINDOWS = [
    ("B", "March",     date(2026, 3, 1),  date(2026, 3, 31)),
    ("C", "April",     date(2026, 4, 1),  date(2026, 4, 30)),
    ("D", "May 1–11",  date(2026, 5, 1),  date(2026, 5, 11)),
]


# ---------------------------------------------------------------------------
# Pulls
# ---------------------------------------------------------------------------

def pull_google_ads(start: date, end: date) -> float:
    client = GoogleAdsClient()
    rows = client.get_campaign_spend(start, end)
    return round(sum(r["spend_gbp"] for r in rows), 2)


def pull_shopify_revenue(start: date, end: date) -> float:
    """
    Sum totalPriceSet across Shopify online-store orders in the window,
    excluding cancelled orders. Uses GraphQL pagination.
    """
    query = (
        'channel:web '
        f'created_at:>={start.strftime("%Y-%m-%dT00:00:00")} '
        f'created_at:<={end.strftime("%Y-%m-%dT23:59:59")}'
    )
    gql = """
    query($cursor: String, $q: String!) {
      orders(first: 100, after: $cursor, query: $q) {
        pageInfo { hasNextPage endCursor }
        edges {
          node {
            cancelledAt
            totalPriceSet { shopMoney { amount } }
          }
        }
      }
    }
    """
    total = 0.0
    cursor = None
    with ShopifyClient() as client:
        while True:
            data = client.execute(gql, {"cursor": cursor, "q": query})
            orders = data.get("orders", {})
            for edge in orders.get("edges", []):
                node = edge["node"]
                if node.get("cancelledAt"):
                    continue
                total += float(node["totalPriceSet"]["shopMoney"]["amount"])
            page = orders.get("pageInfo", {})
            if not page.get("hasNextPage"):
                break
            cursor = page["endCursor"]
    return round(total, 2)


def _amazon_get_with_retry(client, path, params, max_attempts=6):
    """Hit the Orders API with exponential backoff on 429."""
    from scripts.amazon_client import AmazonSPAPIError
    delay = 30.0
    for attempt in range(max_attempts):
        try:
            return client.get(path, params=params)
        except AmazonSPAPIError as e:
            if "429" not in str(e.status if hasattr(e, "status") else e) and "QuotaExceeded" not in str(e):
                raise
            print(f"      429 quota — sleeping {delay:.0f}s (attempt {attempt+1}/{max_attempts})")
            time.sleep(delay)
            delay *= 1.6
    # final try, let exception propagate
    return client.get(path, params=params)


def pull_amazon_revenue(start: date, end: date) -> float:
    """
    Sum OrderTotal across Amazon orders in the window, excluding cancelled.
    Uses /orders/v0/orders with CreatedAfter/CreatedBefore. UK marketplace.
    """
    import os
    from datetime import timedelta
    client = AmazonClient()
    marketplace = os.environ["AMAZON_MARKETPLACE_ID"]
    path = "/orders/v0/orders"

    # SP-API rejects CreatedBefore within 2 minutes of now. If end is today
    # or later, clamp it to (now - 5 minutes) UTC.
    end_dt = datetime.combine(end, datetime.max.time(), tzinfo=timezone.utc)
    now_utc = datetime.now(timezone.utc)
    if end_dt > now_utc - timedelta(minutes=5):
        end_dt = now_utc - timedelta(minutes=5)

    params: dict = {
        "MarketplaceIds":     marketplace,
        "CreatedAfter":       datetime.combine(start, datetime.min.time(),
                                               tzinfo=timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "CreatedBefore":      end_dt.strftime("%Y-%m-%dT%H:%M:%SZ"),
        "OrderStatuses":      "Shipped,Unshipped,PartiallyShipped,InvoiceUnconfirmed,PendingAvailability",
        "MaxResultsPerPage":  100,
    }
    total = 0.0
    count = 0
    next_token: str | None = None
    while True:
        page_params = ({"MarketplaceIds": marketplace, "NextToken": next_token}
                       if next_token else params)
        page = _amazon_get_with_retry(client, path, page_params)
        payload = page.get("payload", {}) or {}
        for order in payload.get("Orders", []) or []:
            t = order.get("OrderTotal") or {}
            try:
                total += float(t.get("Amount", 0) or 0)
                count += 1
            except (TypeError, ValueError):
                continue
        next_token = payload.get("NextToken")
        if not next_token:
            break
        time.sleep(3.0)  # extra spacing to stay under quota
    print(f"    (Amazon: {count} orders summed)")
    return round(total, 2)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    # 1) Backup
    if not BACKUP.exists():
        shutil.copy2(WB_PATH, BACKUP)
        print(f"Backed up to: {BACKUP.name}")
    else:
        print(f"Backup already exists: {BACKUP.name}  (not overwriting)")

    # 2) Pull all values
    results: dict[str, dict[str, float]] = {"google": {}, "shopify": {}, "amazon": {}}
    for col, label, start, end in WINDOWS:
        print(f"\n=== {label} ({start} → {end}) ===")
        print("  Google Ads (Mow) …")
        results["google"][col]  = pull_google_ads(start, end)
        print(f"    £{results['google'][col]:,.2f}")
        print("  Shopify (Mow Website) …")
        results["shopify"][col] = pull_shopify_revenue(start, end)
        print(f"    £{results['shopify'][col]:,.2f}")
        print("  Amazon (FBA Sales) …")
        results["amazon"][col]  = pull_amazon_revenue(start, end)
        print(f"    £{results['amazon'][col]:,.2f}")

    # 3) Write into the workbook
    wb = load_workbook(WB_PATH)
    ws = wb["Sheet1"]

    for col in ("B", "C", "D"):
        ws[f"{col}6"]  = results["google"][col]   # Google Ads (Mow)
        ws[f"{col}22"] = results["shopify"][col]  # Mow Website
        ws[f"{col}26"] = results["amazon"][col]   # Amazon FBA Sales

    wb.save(WB_PATH)

    # 4) Summary
    print("\n" + "=" * 60)
    print("Written to workbook (in place):")
    print(f"  Backup at: {BACKUP.name}")
    print("\n  Row 6  Google Ads (Mow)     Mar   £{:>10,.2f}".format(results["google"]["B"]))
    print("                                Apr   £{:>10,.2f}".format(results["google"]["C"]))
    print("                                May   £{:>10,.2f}".format(results["google"]["D"]))
    print("\n  Row 22 Mow Website          Mar   £{:>10,.2f}".format(results["shopify"]["B"]))
    print("                                Apr   £{:>10,.2f}".format(results["shopify"]["C"]))
    print("                                May   £{:>10,.2f}".format(results["shopify"]["D"]))
    print("\n  Row 26 Amazon FBA Sales     Mar   £{:>10,.2f}".format(results["amazon"]["B"]))
    print("                                Apr   £{:>10,.2f}".format(results["amazon"]["C"]))
    print("                                May   £{:>10,.2f}".format(results["amazon"]["D"]))
    print("\n  Still need YOU to fill:")
    print("    Row 7   Amazon Advertising (need Amazon Ads console login)")
    print("    Row 8   B&Q Advertising")
    print("    Row 11  Email Marketing (Salesfire fee)")
    print("    Row 14  Commissions (Mow — Wayne)")
    print("    Row 15  Commissions (Mow — ClickSlice)")


if __name__ == "__main__":
    main()
