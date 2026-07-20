#!/usr/bin/env python3
"""
Template-first batch CSV generator for the /range-csv-batch skill.

The Range analogue of scripts/tesco_csv_batch.py, built the SAME way: the shape
of a category — its attributes and which are required — comes ENTIRELY from a
downloaded Mirakl category template XLSX dropped into range/templates/. There is
NO call to The Range's API to discover fields; the template is the single source
of truth (see memory/mirakl_template_is_source_of_truth.md). The only live calls
are to Shopify (product copy, images) — never to The Range's schema.

A Range template is a standard Mirakl 3-sheet workbook, like Tesco's:
    Data           row 0 = portal labels, row 1 = api codes in canonical order
    ReferenceData  row 0 = value-list attribute api codes, rows below = values
    Columns        row 0 = (Code, Label, Description, Value example, <breadcrumb>)
                   each row = [api_code, label, description, value_example, marker]
                   marker ∈ {REQUIRED, RECOMMENDED, OPTIONAL}

The Range deltas vs Tesco (confirmed by reading real templates 2026-06-17):
  - Core api codes differ: category/shop_sku/title/brand/gtin/description/
    main_image/image_1..image_20/feature_1..10/variant_group_code
    (supplied by scripts.mirakl_operators.THERANGE.field_schema). Note the EAN
    column is `gtin`, not Tesco's `barcode`.
  - The category VALUE is the breadcrumb STRING itself (e.g.
    "DIY/Power Tools/Chainsaws"), fixed per template — NOT a Shopify gid (Tesco)
    or PIM code (B&Q). So it validates against its own 1-entry value-list
    normally; no gid exemption is needed.
  - MULTI-BRAND catalogue: brand = Shopify vendor per-product (no Spectrum
    default). Colour is per-brand/per-SKU (config/range_overrides.json), not a
    blanket default.
  - Heavier REQUIRED set: colour trio (colour_group value-list + display name +
    hex), feature_1..3, main_product_image (Yes/No) + main_product_name,
    variant_group_code, and on machinery dimensions/material/wattage with
    separate *_uom value-list partner columns.
  - Routing: The Range breadcrumbs are its OWN taxonomy, so SKUs are routed to a
    template via config/range_category_map.json (Shopify category -> template),
    NOT by breadcrumb equality with Shopify (which never matches).

Subcommands (parallel to tesco_csv_batch):
    list-templates                 Scan range/templates/*.xlsx -> categories.
    suggest-template --skus         Route SKUs to templates via the category map.
    init-batch --category --skus    Resolve template, create state + CSV header.
    gather --batch-state --sku      Pull Shopify, re-host images to JPEG, compute
                                    the deterministic Range core, surface spec
                                    columns + Shopify sources for AI enrichment.
    write-row --batch-state --sku --row-json
                                    Merge core + AI spec cells, validate, append.
    finalize --batch-state          Print batch summary.

Run from the repo root. Output -> range/csv-output/. Push the reviewed CSV with
    scripts/mirakl_push_products.py --operator THERANGE --file <csv>
"""
from __future__ import annotations

import argparse
import csv
import html
import io
import json
import os
import re
import sys
import time
import warnings
from datetime import datetime, timezone
from typing import Any

import requests

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
_REPO_ROOT = os.path.dirname(_THIS_DIR)
for _p in (_THIS_DIR, _REPO_ROOT):
    if _p not in sys.path:
        sys.path.insert(0, _p)

from scripts.mirakl_operators import THERANGE          # noqa: E402
from scripts.shopify_client import ShopifyClient        # noqa: E402
from scripts.blog_publish import (                       # noqa: E402
    STAGED_UPLOADS, FILE_CREATE, FILE_POLL,
)
import marketplace_dims                                  # noqa: E402

_TEMPLATES_DIR = os.path.join(_REPO_ROOT, "range/templates")
_OUTPUT_DIR = os.path.join(_REPO_ROOT, "range/csv-output")
_OVERRIDES_PATH = os.path.join(_REPO_ROOT, "config", "range_overrides.json")
_CATEGORY_MAP_PATH = os.path.join(_REPO_ROOT, "config", "range_category_map.json")
_JPG_CACHE_DIR = os.path.join(_REPO_ROOT, "workdir", "mirakl-range")
_JPG_CACHE_PATH = os.path.join(_JPG_CACHE_DIR, "image_jpg_cache.json")

# Operator backbone — core column names, copy scrub, catalogue-wide defaults.
_FS = THERANGE.field_schema
_CP = THERANGE.compliance

_TAG_RE = re.compile(r"<[^>]+>")
_JPG_EXT_RE = re.compile(r"\.jpe?g$", re.IGNORECASE)

# Range templates use REQUIRED/RECOMMENDED/OPTIONAL. The category marker column
# in the Columns sheet starts at index 4 (after Code/Label/Description/Example).
_CATEGORY_COL_MIN = 4
_REQUIRED_MARKERS = {"REQUIRED", "RECOMMENDED", "OPTIONAL"}

# Shopify weight unit -> Range *_uom value-list label.
_SHOPIFY_WEIGHT_UOM = {
    "KILOGRAMS": "Kilograms", "GRAMS": "Grams",
    "POUNDS": "Pounds", "OUNCES": "Ounces",
}


# ---------------------------------------------------------------------------
# Time + state helpers
# ---------------------------------------------------------------------------

def _utcnow_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M")


def _slugify(s: str) -> str:
    s = re.sub(r"[^a-zA-Z0-9_-]+", "-", s.strip().lower())
    return re.sub(r"-+", "-", s).strip("-") or "batch"


def _safe_sku(sku: str) -> str:
    """SKU as a flat filename component — some SKUs contain '/' (e.g.
    '2T2010483/M25') which would otherwise be treated as a path separator and
    scatter context/row JSON into subdirectories."""
    return re.sub(r"[/\\]+", "_", sku)


def _load_state(path: str) -> dict[str, Any]:
    with open(path) as fh:
        return json.load(fh)


def _save_state(path: str, state: dict[str, Any]) -> None:
    with open(path, "w") as fh:
        json.dump(state, fh, indent=2, ensure_ascii=False)
        fh.write("\n")


# ---------------------------------------------------------------------------
# Overrides + category routing map (config/)
# ---------------------------------------------------------------------------

def _load_overrides() -> dict[str, Any]:
    try:
        with open(_OVERRIDES_PATH) as fh:
            return json.load(fh)
    except FileNotFoundError:
        return {}


def _load_category_map() -> dict[str, str]:
    """Shopify-category -> Range-template-leaf routing map. Keys are normalised
    Shopify category breadcrumbs or leaf segments; values are Range template
    leaves (e.g. 'Cordless Lawn Mowers')."""
    try:
        with open(_CATEGORY_MAP_PATH) as fh:
            data = json.load(fh)
    except FileNotFoundError:
        return {}
    return {k: v for k, v in data.items() if not k.startswith("_")}


# Sentinel map value: a Shopify mower category that The Range splits by power
# source into Cordless/Electric/Petrol Lawn Mowers.
_MOWER_BY_POWER = "__MOWER_BY_POWER__"
_POWER_TO_MOWER_LEAF = {
    "cordless": "Cordless Lawn Mowers",
    "electric": "Electric Lawn Mowers",
    "petrol":   "Petrol Lawn Mowers",
}


def _detect_power(product: dict[str, Any]) -> str | None:
    """Infer power source from title/type/tags/metafields. Returns
    'cordless' | 'electric' | 'petrol' | None."""
    hay = " ".join([
        product.get("title") or "",
        product.get("productType") or "",
        " ".join(product.get("tags") or []),
    ]).lower()
    for e in product.get("metafields", {}).get("edges", []):
        n = e["node"]
        if any(t in (n.get("key") or "").lower() for t in ("power", "fuel", "engine")):
            hay += " " + str(n.get("value") or "").lower()
    if any(t in hay for t in ("petrol", "diesel", " cc", "engine", "stroke")):
        return "petrol"
    if any(t in hay for t in ("cordless", "battery", "40v", "20v", "18v", "36v", "lithium")):
        return "cordless"
    if any(t in hay for t in ("electric", "corded", "mains", "230v", "240v", "plug-in")):
        return "electric"
    return None


def _override_colour(ov: dict, sku: str, vendor: str | None) -> str | None:
    """colour_group: per-SKU wins over per-brand."""
    sku_colour = ov.get("sku_colour", {})
    if sku in sku_colour:
        return sku_colour[sku]
    brand_colour = ov.get("brand_colour", {})
    if vendor and vendor.lower() in brand_colour:
        return brand_colour[vendor.lower()]
    return None


# ---------------------------------------------------------------------------
# Template reading — THE SOURCE OF TRUTH for category shape (no API)
# ---------------------------------------------------------------------------

def _open_template(path: str):
    import openpyxl

    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def _category_col_index(header: tuple) -> int | None:
    """The category marker column is the first cell at index >= 4 with a
    non-empty header (the breadcrumb). Range templates carry exactly one."""
    for i in range(_CATEGORY_COL_MIN, len(header)):
        if header[i]:
            return i
    return None


def _list_templates() -> list[dict[str, Any]]:
    """Scan range/templates/*.xlsx -> one category per file.

    Returns [{file, category (breadcrumb), leaf, category_col_index}]."""
    out: list[dict[str, Any]] = []
    if not os.path.isdir(_TEMPLATES_DIR):
        return out
    for name in sorted(os.listdir(_TEMPLATES_DIR)):
        if not name.endswith(".xlsx") or name.startswith("~"):
            continue
        path = os.path.join(_TEMPLATES_DIR, name)
        wb = _open_template(path)
        try:
            header = next(wb["Columns"].iter_rows(values_only=True))
            idx = _category_col_index(header)
            if idx is None:
                continue
            breadcrumb = header[idx]
            out.append({
                "file": os.path.relpath(path, _REPO_ROOT),
                "category": breadcrumb,
                "leaf": str(breadcrumb).split("/")[-1],
                "category_col_index": idx,
            })
        finally:
            wb.close()
    return out


def _resolve_category(query: str) -> tuple[str, int, str]:
    """Find which template file holds `query`. Match (case-insensitive) against
    the full breadcrumb or its leaf segment; exact-leaf wins ties.

    Returns (template_path, category_col_index, breadcrumb)."""
    matches: list[tuple[str, int, str, bool]] = []
    q = query.strip().lower()
    for t in _list_templates():
        path = os.path.join(_REPO_ROOT, t["file"])
        bc = str(t["category"])
        leaf = t["leaf"].lower()
        exact_leaf = leaf == q
        if exact_leaf or q in bc.lower():
            matches.append((path, t["category_col_index"], bc, exact_leaf))
    if not matches:
        raise SystemExit(
            f"No category matches '{query}' in any range/templates/*.xlsx. "
            f"Run `list-templates` to see what's available."
        )
    exact = [m for m in matches if m[3]]
    if exact:
        p, i, bc, _ = exact[0]
        return p, i, bc
    if len(matches) > 1:
        raise SystemExit(
            f"Ambiguous category '{query}' matched: {[m[2] for m in matches]}. "
            f"Be more specific."
        )
    p, i, bc, _ = matches[0]
    return p, i, bc


def _extract_column_meta(template_path: str, cat_col_index: int) -> list[dict[str, Any]]:
    """Read the template -> per-column metadata in Data-sheet order:
        {api_code, portal_label, description, required, value_list, unit, numeric_only}
    `required` is the REQUIRED/RECOMMENDED/OPTIONAL marker for this category."""
    wb = _open_template(template_path)
    try:
        rows = list(wb["Columns"].iter_rows(values_only=True))
        col_entries: dict[str, dict[str, Any]] = {}
        for row in rows[1:]:
            api_code = row[0]
            if not api_code:
                continue
            col_entries[api_code] = {
                "api_code": api_code,
                "portal_label": row[1],
                "description": row[2] or "",
                "required": row[cat_col_index] if cat_col_index < len(row) else None,
            }

        data_rows = list(wb["Data"].iter_rows(values_only=True))
        api_codes_in_order = [c for c in data_rows[1] if c]

        # ReferenceData -> value lists. Like Tesco, col 0 (category) IS a genuine
        # value list — do NOT skip it.
        ref_rows = list(wb["ReferenceData"].iter_rows(values_only=True))
        ref_header = ref_rows[0]
        value_lists: dict[str, list[str]] = {}
        for col_idx, api_code in enumerate(ref_header):
            if not api_code:
                continue
            values = [str(r[col_idx]) for r in ref_rows[1:]
                      if col_idx < len(r) and r[col_idx] not in (None, "")]
            if values:
                value_lists[api_code] = values

        out: list[dict[str, Any]] = []
        for api_code in api_codes_in_order:
            entry = col_entries.get(api_code, {
                "api_code": api_code, "portal_label": api_code,
                "description": "", "required": None,
            })
            label = entry["portal_label"] or ""
            desc = entry["description"]
            unit_match = re.search(r"\(([a-zA-Z°²³µ/]+)\)\s*$", label)
            unit = unit_match.group(1) if unit_match else None
            entry.update({
                "value_list": value_lists.get(api_code),
                "unit": unit,
                "numeric_only": "numeric" in desc.lower() and "only" in desc.lower(),
            })
            out.append(entry)
        return out
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Shopify pull
# ---------------------------------------------------------------------------

def _shopify_pull_full(sku: str, ean: str | None = None) -> tuple[dict[str, Any] | None, str | None]:
    """Resolve a Shopify variant. Try the SKU first; if not found and an EAN is
    given, fall back to a barcode lookup (the SKU-mismatch case). Returns
    (variant, matched_by) where matched_by is 'sku' | 'ean' | None."""
    query = """
    query($q: String!) {
      productVariants(first: 5, query: $q) {
        edges {
          node {
            sku
            barcode
            price
            inventoryQuantity
            inventoryItem { measurement { weight { value unit } } }
            metafields(first: 100) { edges { node { namespace key type value } } }
            product {
              id
              title
              descriptionHtml
              vendor
              productType
              tags
              status
              featuredImage { url altText }
              images(first: 20) { edges { node { url altText } } }
              category { id name fullName }
              metafields(first: 100) { edges { node { namespace key type value } } }
            }
          }
        }
      }
    }
    """
    with ShopifyClient() as client:
        res = client.execute(query, variables={"q": f"sku:{sku}"})
        for e in res["productVariants"]["edges"]:
            if e["node"]["sku"] == sku:
                return e["node"], "sku"
        if ean:
            res = client.execute(query, variables={"q": f"barcode:{ean}"})
            edges = res["productVariants"]["edges"]
            if edges:
                return edges[0]["node"], "ean"
    return None, None


def _shopify_deref_metaobjects(gids: list[str]) -> dict[str, str]:
    if not gids:
        return {}
    query = """
    query($ids: [ID!]!) {
      nodes(ids: $ids) {
        ... on Metaobject { id displayName handle fields { key value } }
      }
    }
    """
    out: dict[str, str] = {}
    with ShopifyClient() as client:
        for i in range(0, len(gids), 50):
            res = client.execute(query, variables={"ids": gids[i:i + 50]})
            for node in res["nodes"]:
                if not node:
                    continue
                display = node.get("displayName") or node.get("handle") or node["id"]
                for f in node.get("fields") or []:
                    if f["key"] in ("label", "name") and f.get("value"):
                        display = f["value"]
                        break
                out[node["id"]] = display
    return out


def _flatten_metafields(variant: dict[str, Any], gid_lookup: dict[str, str]) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    for owner_label, mf_block in (
        ("PRODUCTVARIANT", variant.get("metafields", {}).get("edges", [])),
        ("PRODUCT", variant.get("product", {}).get("metafields", {}).get("edges", [])),
    ):
        for edge in mf_block:
            mf = edge["node"]
            key = f"{mf['namespace']}.{mf['key']}"
            raw, mtype = mf["value"], mf["type"]
            derefed: Any = raw
            if mtype == "list.metaobject_reference" and raw:
                try:
                    derefed = [gid_lookup.get(g, g) for g in json.loads(raw)]
                except (json.JSONDecodeError, TypeError):
                    pass
            elif mtype == "metaobject_reference" and raw:
                derefed = gid_lookup.get(raw, raw)
            elif mtype == "json" and raw:
                try:
                    derefed = json.loads(raw)
                except (json.JSONDecodeError, TypeError):
                    pass
            out[key] = {"owner_type": owner_label, "type": mtype,
                        "raw_value": raw, "value": derefed}
    return out


# ---------------------------------------------------------------------------
# Image re-host — JPEG is universally accepted; re-host guarantees acceptance
# ---------------------------------------------------------------------------

def _is_jpg_url(url: str) -> bool:
    if not url:
        return False
    return bool(_JPG_EXT_RE.search(url.split("?", 1)[0]))


def _load_jpg_cache() -> dict:
    if os.path.exists(_JPG_CACHE_PATH):
        try:
            with open(_JPG_CACHE_PATH) as fh:
                return json.load(fh)
        except (OSError, json.JSONDecodeError):
            return {}
    return {}


def _save_jpg_cache(cache: dict) -> None:
    os.makedirs(_JPG_CACHE_DIR, exist_ok=True)
    with open(_JPG_CACHE_PATH, "w") as fh:
        json.dump(cache, fh, indent=2, sort_keys=True)


def _to_jpeg_bytes(raw: bytes) -> bytes:
    from PIL import Image
    im = Image.open(io.BytesIO(raw))
    if im.mode in ("RGBA", "LA") or (im.mode == "P" and "transparency" in im.info):
        im = im.convert("RGBA")
        bg = Image.new("RGB", im.size, (255, 255, 255))
        bg.paste(im, mask=im.split()[-1])
        im = bg
    else:
        im = im.convert("RGB")
    buf = io.BytesIO()
    im.save(buf, "JPEG", quality=88, progressive=True, optimize=True)
    return buf.getvalue()


def _upload_jpg(client, filename: str, data: bytes) -> str:
    staged = client.execute(STAGED_UPLOADS, {"input": [{
        "filename": filename, "mimeType": "image/jpeg",
        "resource": "FILE", "httpMethod": "POST",
    }]})
    target = staged["stagedUploadsCreate"]["stagedTargets"][0]
    form = {p["name"]: p["value"] for p in target["parameters"]}
    up = requests.post(target["url"], data=form,
                       files={"file": (filename, data, "image/jpeg")}, timeout=120)
    up.raise_for_status()
    created = client.execute(FILE_CREATE, {"files": [{
        "originalSource": target["resourceUrl"], "contentType": "IMAGE", "alt": "",
    }]})
    errs = created["fileCreate"]["userErrors"]
    if errs:
        raise RuntimeError(f"fileCreate: {errs}")
    node = created["fileCreate"]["files"][0]
    fid = node["id"]
    url = (node.get("image") or {}).get("url")
    for _ in range(30):
        if url:
            return url
        time.sleep(2)
        n = client.execute(FILE_POLL, {"id": fid})["node"] or {}
        if n.get("fileStatus") == "READY":
            url = (n.get("image") or {}).get("url")
        elif n.get("fileStatus") == "FAILED":
            raise RuntimeError("Shopify fileStatus FAILED")
    raise RuntimeError("timed out waiting for Shopify to process the jpg")


def _ensure_jpg(client, url: str, cache: dict) -> str | None:
    if _is_jpg_url(url):
        return url
    key = url.split("?", 1)[0]
    if key in cache:
        return cache[key]
    try:
        raw = requests.get(url, headers={"Accept": "*/*"}, timeout=60)
        raw.raise_for_status()
        data = _to_jpeg_bytes(raw.content)
        base = re.sub(r"[^A-Za-z0-9_-]", "-", key.split("/")[-1].rsplit(".", 1)[0])
        new_url = _upload_jpg(client, f"range-{base}.jpg", data)
    except Exception as e:  # noqa: BLE001
        print(f"    [rehost failed] {key.split('/')[-1]}: {e}", file=sys.stderr)
        return None
    cache[key] = new_url
    _save_jpg_cache(cache)
    return new_url


# ---------------------------------------------------------------------------
# Copy helpers
# ---------------------------------------------------------------------------

def _strip_html(s: str | None) -> str:
    if not s:
        return ""
    return " ".join(html.unescape(_TAG_RE.sub(" ", s)).split())


def _norm_breadcrumb(s: str | None) -> str:
    if not s:
        return ""
    parts = re.split(r"\s*[/>]\s*", s)
    return "/".join(p.strip().lower() for p in parts if p.strip())


def _product_images(product: dict[str, Any]) -> list[str]:
    urls: list[str] = []
    fi = product.get("featuredImage")
    if fi and fi.get("url"):
        urls.append(fi["url"])
    for e in product.get("images", {}).get("edges", []):
        u = (e.get("node") or {}).get("url")
        if u and u not in urls:
            urls.append(u)
    return urls


# ---------------------------------------------------------------------------
# Deterministic Range core (the mechanical, non-AI columns)
# ---------------------------------------------------------------------------

def _deterministic_core(variant: dict[str, Any], breadcrumb: str,
                        ov: dict[str, Any], client, valid_codes: set[str],
                        out_sku: str | None = None,
                        ean: str | None = None) -> tuple[dict[str, str], list[str], dict]:
    """Compute the columns we can fill mechanically: category breadcrumb, sku,
    gtin, scrubbed title/description, brand, colour trio, variant group, main-
    product flags, Shopify weight, catalogue defaults, and the JPEG-re-hosted
    images. Only keys that are real columns in this template (valid_codes)
    survive. Returns (core, notes, scrub_hits)."""
    product = variant["product"]
    out_sku = out_sku or variant["sku"]
    notes: list[str] = []
    scrub_hits: dict[str, list[str]] = {}

    title_clean = _CP.clean_name(product.get("title") or "")
    body_clean, body_hits = _CP.sanitise_prose(_strip_html(product.get("descriptionHtml")))
    _, name_hits = _CP.scrub(product.get("title") or "")
    hits = sorted(set(body_hits) | set(name_hits))
    if hits:
        scrub_hits[out_sku] = hits

    core: dict[str, str] = {
        _FS.category: breadcrumb,
        _FS.sku: out_sku,
        _FS.ean: (ean or variant.get("barcode") or "").strip(),
        _FS.name: title_clean,
        _FS.body: body_clean,
        # Variant grouping: standalone product -> its own SKU is the group code,
        # it is the main product, and its name/image represent the group.
        _FS.variant_group: out_sku,
        "main_product_name": title_clean,
    }

    # Brand — Shopify vendor (The Range is multi-brand; no Spectrum default).
    vendor = (product.get("vendor") or "").strip()
    if vendor:
        core["brand"] = vendor

    # Colour trio — per-SKU / per-brand override (REQUIRED but value-list).
    colour = _override_colour(ov, out_sku, vendor)
    if colour:
        core["colour_group"] = colour
        core["colour_display_name"] = colour
        hexmap = ov.get("colour_hex", {})
        if colour in hexmap:
            core["colour_hex_code"] = hexmap[colour]
    else:
        notes.append(
            f"colour_group unresolved for {out_sku} (brand '{vendor}') — REQUIRED. "
            f"Add a brand_colour/sku_colour entry in config/range_overrides.json."
        )

    # Catalogue-wide value-list defaults (made_to_order, main_product_image, …).
    defaults = {**THERANGE.common_attributes, **ov.get("defaults", {})}
    for k, v in defaults.items():
        if v is not None and k not in core:
            core[k] = str(v)

    # GPSR economic-operator block (only the filled ones).
    for k, v in (ov.get("gpsr", {}) or {}).items():
        if k.startswith("_") or not v:
            continue
        core[k] = str(v)

    # Shopify weight -> the *first* weight column present in this template
    # (weight_61 on machinery; absent on simple templates like Chainsaws).
    w = (variant.get("inventoryItem") or {}).get("measurement", {}).get("weight")
    if w and w.get("value"):
        weight_code = next((c for c in valid_codes
                            if c.startswith("weight_") and not c.endswith("_uom")), None)
        if weight_code:
            core[weight_code] = f"{float(w['value']):.2f}"
            uom_code = f"{weight_code}_uom"
            uom = _SHOPIFY_WEIGHT_UOM.get((w.get("unit") or "").upper())
            if uom and uom_code in valid_codes:
                core[uom_code] = uom

    # Images -> JPEG, in order, into main_image + image_1..image_N.
    cache = _load_jpg_cache()
    slots = [_FS.image_main, *_FS.extra_images]
    img_urls = _product_images(product)
    jpgs: list[str] = []
    for u in img_urls:
        ju = _ensure_jpg(client, u, cache)
        if ju:
            jpgs.append(ju)
    if len(jpgs) < len(img_urls):
        notes.append(f"{len(img_urls) - len(jpgs)} image(s) failed JPEG re-host")
    for slot, ju in zip(slots, jpgs):
        core[slot] = ju
    if not jpgs:
        notes.append("no usable images (main_image is REQUIRED) — fix before push")

    # Drop anything that isn't a real column in this template.
    core = {k: v for k, v in core.items() if k in valid_codes and v not in (None, "")}
    return core, notes, scrub_hits


# ---------------------------------------------------------------------------
# Validator
# ---------------------------------------------------------------------------

def _canonicalize(value: str, column_meta: dict[str, Any]) -> str:
    """If a value-list value matches an allowed entry only by case, return the
    list's canonical casing. The Range value-lists are case-sensitive."""
    vl = column_meta.get("value_list")
    if not vl or value in (None, ""):
        return value
    s = str(value)
    if s in vl:
        return s
    low = s.lower()
    for entry in vl:
        if entry.lower() == low:
            return entry
    return value


def _validate_cell(value: str, column_meta: dict[str, Any]) -> tuple[bool, str | None]:
    if value is None or value == "":
        return True, None
    s = str(value)
    api_code = column_meta["api_code"]
    if column_meta.get("value_list"):
        if s not in column_meta["value_list"]:
            lower = s.lower()
            near = [v for v in column_meta["value_list"]
                    if lower in v.lower() or v.lower() in lower][:3]
            hint = f"; nearest in value list: {near}" if near else ""
            return False, f"'{api_code}': value '{s}' not in value list{hint}"
    if column_meta.get("numeric_only"):
        try:
            float(s)
        except ValueError:
            return False, f"'{api_code}' expects numeric only; got '{s}'"
    return True, None


# ---------------------------------------------------------------------------
# Subcommands
# ---------------------------------------------------------------------------

def _cmd_list_templates() -> int:
    templates = _list_templates()
    print(json.dumps({
        "templates": templates,
        "categories": [{"category": t["category"], "leaf": t["leaf"],
                        "template": t["file"]} for t in templates],
    }, indent=2, ensure_ascii=False))
    return 0


def _cmd_suggest_template(skus: list[str]) -> int:
    """Route SKUs to Range templates. The Range breadcrumbs are its own taxonomy
    (never match Shopify), so we read each product's Shopify category and map it
    via config/range_category_map.json (Shopify category -> Range template leaf).

    Each item may be `SKU` or `SKU=EAN` (barcode fallback for SKU-mismatch).
    Reports `suggestions`, `unmatched_skus` (Shopify category has no map entry —
    add one), and `skus_not_found`."""
    templates = _list_templates()
    by_leaf = {t["leaf"].lower(): t for t in templates}
    cat_map = _load_category_map()
    norm_map = {_norm_breadcrumb(k): v for k, v in cat_map.items()}

    query = """
    query($q: String!) {
      productVariants(first: 5, query: $q) {
        edges { node { sku product {
          title productType tags
          category { id name fullName }
          metafields(first: 50) { edges { node { key value } } }
        } } }
      }
    }
    """
    suggestions: dict[str, dict[str, Any]] = {}
    unmatched: list[dict[str, Any]] = []
    not_found: list[str] = []
    with ShopifyClient() as client:
        for item in skus:
            sku, _, ean = item.partition("=")
            sku = sku.strip(); ean = ean.strip()
            res = client.execute(query, variables={"q": f"sku:{sku}"})
            hit = next((e["node"] for e in res["productVariants"]["edges"]
                        if e["node"]["sku"] == sku), None)
            if hit is None and ean:
                res = client.execute(query, variables={"q": f"barcode:{ean}"})
                edges = res["productVariants"]["edges"]
                hit = edges[0]["node"] if edges else None
            if hit is None:
                not_found.append(sku)
                continue
            cat = hit["product"].get("category") or {}
            full = cat.get("fullName")
            # map by full breadcrumb, then by leaf segment
            target_leaf = norm_map.get(_norm_breadcrumb(full))
            if not target_leaf and full:
                leaf = full.split(" > ")[-1].split("/")[-1].strip()
                target_leaf = norm_map.get(_norm_breadcrumb(leaf))
            # Mowers: resolve the power-source split to a concrete template leaf.
            if target_leaf == _MOWER_BY_POWER:
                power = _detect_power(hit["product"])
                target_leaf = _POWER_TO_MOWER_LEAF.get(power or "")
            t = by_leaf.get((target_leaf or "").lower()) if target_leaf else None
            if t:
                s = suggestions.setdefault(t["file"], {
                    "template": t["file"], "category": t["category"],
                    "leaf": t["leaf"], "skus": [],
                })
                s["skus"].append(sku)
            else:
                unmatched.append({
                    "sku": sku,
                    "shopify_category": full,
                    "mapped_to": target_leaf,
                    "reason": ("no config/range_category_map.json entry maps this "
                               "Shopify category to a Range template (or the mapped "
                               "leaf has no template). Add a mapping."),
                })

    sugg_list = list(suggestions.values())
    print(json.dumps({
        "suggestions": sugg_list,
        "unmatched_skus": unmatched,
        "skus_not_found": not_found,
        "single_template": len(sugg_list) == 1 and not unmatched and not not_found,
    }, indent=2, ensure_ascii=False))
    return 0


def _cmd_init_batch(category: str, skus: list[str], batch_name: str | None) -> int:
    template_path, col_idx, breadcrumb = _resolve_category(category)
    stamp = _stamp()
    name = batch_name or _slugify(str(breadcrumb).split("/")[-1])
    batch_id = f"{name}_{stamp}"

    state_path = os.path.join(_OUTPUT_DIR, f".state_{batch_id}.json")
    csv_path = os.path.join(_OUTPUT_DIR, f"{batch_id}.csv")
    log_path = os.path.join(_OUTPUT_DIR, f"{batch_id}.log.md")

    columns = _extract_column_meta(template_path, col_idx)
    api_codes = [c["api_code"] for c in columns]
    required = [c["api_code"] for c in columns if c["required"] == "REQUIRED"]

    state = {
        "batch_id": batch_id,
        "batch_name": name,
        "created_at": _utcnow_iso(),
        "category_breadcrumb": breadcrumb,
        "template_file": os.path.relpath(template_path, _REPO_ROOT),
        "category_col_index": col_idx,
        "csv_path": os.path.relpath(csv_path, _REPO_ROOT),
        "log_path": os.path.relpath(log_path, _REPO_ROOT),
        "skus": skus,
        "processed_skus": [],
        "stats": {"cells_filled": 0, "cells_unknown": 0, "validator_failures": 0},
        "column_count": len(columns),
        "required_columns": required,
    }
    os.makedirs(_OUTPUT_DIR, exist_ok=True)
    _save_state(state_path, state)

    with open(csv_path, "w", newline="", encoding="utf-8") as fh:
        csv.writer(fh, delimiter=";", quoting=csv.QUOTE_ALL).writerow(api_codes)

    with open(log_path, "w") as fh:
        fh.write(f"# Range CSV Batch — {name}, {stamp}\n")
        fh.write(f"Template: {os.path.relpath(template_path, _REPO_ROOT)}\n")
        fh.write(f"Category: {breadcrumb}\n")
        fh.write(f"SKUs: {len(skus)}\n")
        fh.write(f"Columns: {len(columns)} ({len(required)} REQUIRED)\n\n")

    print(json.dumps({
        "ok": True, "batch_id": batch_id,
        "state_path": os.path.relpath(state_path, _REPO_ROOT),
        "csv_path": state["csv_path"], "log_path": state["log_path"],
        "template": state["template_file"], "category": breadcrumb,
        "column_count": len(columns), "required_count": len(required),
    }, indent=2, ensure_ascii=False))
    return 0


def _cmd_gather(state_path: str, sku: str, ean: str | None = None) -> int:
    state = _load_state(state_path)
    ov = _load_overrides()

    if sku in ov.get("exclude_skus", {}):
        print(json.dumps({
            "ok": False, "sku": sku, "excluded": True,
            "reason": ov["exclude_skus"][sku],
        }, indent=2))
        return 0

    variant, matched_by = _shopify_pull_full(sku, ean)
    if variant is None:
        hint = "" if ean else " (no EAN supplied for a barcode fallback)"
        print(json.dumps({"error": f"SKU {sku} not found on Shopify{hint}"}), file=sys.stderr)
        return 2

    product = variant["product"]
    warnings_out: list[str] = []
    if matched_by == "ean":
        warnings_out.append(
            f"SKU {sku} not on Shopify by SKU — matched by EAN {ean} "
            f"(Shopify SKU '{variant.get('sku')}'). Range shop_sku keeps '{sku}'."
        )

    breadcrumb = state["category_breadcrumb"]

    # --- Deref metaobjects, flatten metafields ---
    all_mf = [e["node"] for e in variant.get("metafields", {}).get("edges", [])]
    all_mf += [e["node"] for e in product.get("metafields", {}).get("edges", [])]
    gids: list[str] = []
    for mf in all_mf:
        if mf["type"] == "metaobject_reference" and mf["value"]:
            gids.append(mf["value"])
        elif mf["type"] == "list.metaobject_reference" and mf["value"]:
            try:
                gids.extend(json.loads(mf["value"]))
            except (json.JSONDecodeError, TypeError):
                pass
    metafields = _flatten_metafields(variant, _shopify_deref_metaobjects(gids))

    # --- Template columns ---
    template_path = os.path.join(_REPO_ROOT, state["template_file"])
    columns = _extract_column_meta(template_path, state["category_col_index"])
    valid_codes = {c["api_code"] for c in columns}

    # --- Deterministic Range core (incl. JPEG re-host) ---
    with ShopifyClient() as client:
        core, core_notes, scrub_hits = _deterministic_core(
            variant, breadcrumb, ov, client, valid_codes, out_sku=sku, ean=ean)
    warnings_out.extend(core_notes)

    auto_codes = {k for k, v in core.items() if v}
    spec_columns = [c for c in columns if c["api_code"] not in auto_codes]

    shopify_summary = {
        "sku": variant["sku"],
        "barcode": variant.get("barcode"),
        "price": variant.get("price"),
        "inventoryQuantity": variant.get("inventoryQuantity"),
        "weight": (variant.get("inventoryItem") or {}).get("measurement", {}).get("weight"),
        "product": {
            "title": product["title"],
            "descriptionHtml": product["descriptionHtml"],
            "vendor": product["vendor"],
            "productType": product["productType"],
            "tags": product["tags"],
            "status": product["status"],
            "category": product.get("category"),
            "images": [e["node"] for e in product["images"]["edges"]],
        },
        "metafields": metafields,
    }

    if (variant.get("inventoryQuantity") or 0) <= 0:
        warnings_out.append(f"SKU {sku} inventoryQuantity={variant.get('inventoryQuantity')} (out of stock)")

    context = {
        "sku": sku,
        "batch_id": state["batch_id"],
        "category_breadcrumb": breadcrumb,
        "auto_filled": core,
        "scrub_hits": scrub_hits.get(sku, []),
        "spec_columns": spec_columns,
        "shopify": shopify_summary,
        "row_output_schema": {
            "filled":  "{<api_code>: <value as string>, ...} — ONLY the spec_columns "
                       "you can support from Shopify data; the auto_filled core is "
                       "already handled, do not repeat it.",
            "unknown": "[{column: <api_code>, reason: <short>}, ...] for spec_columns "
                       "you cannot fill.",
        },
        "validator_will_check": {
            "value_list_fields": "value must appear in the column's value_list (case sensitive)",
            "numeric_fields": "value must parse as a number",
            "uom_fields": "columns ending _uom are value-lists; pair each numeric spec "
                          "(e.g. width_4) with its unit (width_4_uom)",
        },
        "guidance": (
            "Fill REQUIRED spec columns first (dimensions+_uom, material, wattage, "
            "feature_1..3), then RECOMMENDED/OPTIONAL. Map custom.feature_bullets "
            "<li> items to feature_1..N. For each numeric dimension also set its "
            "*_uom partner from the unit value-list. For value-list columns write "
            "an EXACT value from value_list or mark UNKNOWN."
        ),
        "warnings": warnings_out,
    }

    context_path = os.path.join(_OUTPUT_DIR, f"{state['batch_id']}_{_safe_sku(sku)}_context.json")
    with open(context_path, "w") as fh:
        json.dump(context, fh, indent=2, ensure_ascii=False)

    print(json.dumps({
        "ok": True, "sku": sku,
        "context_path": os.path.relpath(context_path, _REPO_ROOT),
        "auto_filled_count": len(auto_codes),
        "spec_columns_count": len(spec_columns),
        "required_uncovered": [c["api_code"] for c in spec_columns if c["required"] == "REQUIRED"],
        "warnings": warnings_out,
    }, indent=2, ensure_ascii=False))
    return 0


def _cmd_write_row(state_path: str, sku: str, row_json_path: str) -> int:
    state = _load_state(state_path)
    if sku in state["processed_skus"]:
        print(json.dumps({"error": f"SKU {sku} already processed"}), file=sys.stderr)
        return 2

    with open(row_json_path) as fh:
        row = json.load(fh)
    ai_filled: dict[str, str] = row.get("filled") or {}
    unknown: list[dict[str, str]] = row.get("unknown") or []

    auto_filled: dict[str, str] = row.get("auto_filled") or {}
    if not auto_filled:
        context_path = os.path.join(_OUTPUT_DIR, f"{state['batch_id']}_{_safe_sku(sku)}_context.json")
        if os.path.exists(context_path):
            with open(context_path) as fh:
                auto_filled = (json.load(fh) or {}).get("auto_filled") or {}

    template_path = os.path.join(_REPO_ROOT, state["template_file"])
    columns = _extract_column_meta(template_path, state["category_col_index"])
    col_meta = {c["api_code"]: c for c in columns}
    api_codes_in_order = [c["api_code"] for c in columns]
    required_codes = {c["api_code"] for c in columns if c["required"] == "REQUIRED"}
    recommended_codes = {c["api_code"] for c in columns if c["required"] == "RECOMMENDED"}

    merged: dict[str, str] = {}
    validator_failures: list[dict[str, str]] = []
    for source in (auto_filled, ai_filled):
        for api_code, value in source.items():
            if api_code not in col_meta:
                validator_failures.append({"column": api_code, "reason": "not in this category's template"})
                continue
            if api_code in merged and merged[api_code]:
                continue
            value = _canonicalize(value, col_meta[api_code])
            passes, reason = _validate_cell(value, col_meta[api_code])
            if not passes:
                validator_failures.append({"column": api_code, "reason": reason, "rejected_value": str(value)})
                unknown.append({"column": api_code, "reason": f"VALIDATOR: {reason}"})
            else:
                merged[api_code] = str(value) if value is not None else ""

    out_row = [merged.get(code, "") for code in api_codes_in_order]
    csv_path = os.path.join(_REPO_ROOT, state["csv_path"])
    with open(csv_path, "a", newline="", encoding="utf-8") as fh:
        csv.writer(fh, delimiter=";", quoting=csv.QUOTE_ALL).writerow(out_row)

    for code in required_codes:
        if not merged.get(code):
            if not any(u["column"] == code for u in unknown):
                unknown.append({"column": code, "reason": "REQUIRED but blank"})

    # Physical product/shipping dimensions blank → surface as gaps regardless of
    # the template's REQUIRED marker (they must never regress silently).
    dimension_gaps = marketplace_dims.missing_dimension_gaps(columns, merged)
    for gap in dimension_gaps:
        if not any(u["column"] == gap["column"] for u in unknown):
            unknown.append(gap)
    if dimension_gaps:
        print(f"[DIMENSION] {sku}: {len(dimension_gaps)} physical dimension "
              f"cell(s) blank — {', '.join(g['column'] for g in dimension_gaps)}",
              file=sys.stderr)

    cells_filled = sum(1 for v in merged.values() if v)
    log_path = os.path.join(_REPO_ROOT, state["log_path"])
    with open(log_path, "a") as fh:
        fh.write(f"\n## {sku}\n")
        fh.write(f"Filled: {cells_filled} cells   Unknown: {len(unknown)} cells   "
                 f"Validator failures: {len(validator_failures)}\n\n")
        if unknown:
            fh.write("### Gaps (cells still blank)\n")
            for u in unknown:
                col = u["column"]
                tag = ("[REQUIRED]" if col in required_codes
                       else "[RECOMMENDED]" if col in recommended_codes else "[OPTIONAL]")
                if str(u.get("reason", "")).startswith("VALIDATOR:"):
                    tag = "[VALIDATOR]"
                if u.get("dimension"):
                    tag = "[DIMENSION]"
                fh.write(f"- {tag} {col} — {u.get('reason') or 'no reason given'}\n")
            fh.write("\n")

    state["processed_skus"].append(sku)
    state["stats"]["cells_filled"] += cells_filled
    state["stats"]["cells_unknown"] += len(unknown)
    state["stats"]["validator_failures"] += len(validator_failures)
    state["stats"].setdefault("dimensions_blank", 0)
    state["stats"]["dimensions_blank"] += len(dimension_gaps)
    _save_state(state_path, state)

    print(json.dumps({
        "ok": True, "sku": sku,
        "cells_filled": cells_filled,
        "cells_unknown": len(unknown),
        "validator_failures": validator_failures,
        "required_blank": [c for c in required_codes if not merged.get(c)],
        "dimensions_blank": [g["column"] for g in dimension_gaps],
        "progress": f"{len(state['processed_skus'])}/{len(state['skus'])}",
    }, indent=2, ensure_ascii=False))
    return 0


def _cmd_finalize(state_path: str) -> int:
    state = _load_state(state_path)
    missing = [s for s in state["skus"] if s not in state["processed_skus"]]
    print(json.dumps({
        "ok": True, "batch_id": state["batch_id"],
        "csv_path": state["csv_path"], "log_path": state["log_path"],
        "skus_processed": len(state["processed_skus"]),
        "skus_missing": missing, "stats": state["stats"],
    }, indent=2, ensure_ascii=False))
    return 0


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(prog="range_csv_batch")
    sub = parser.add_subparsers(dest="cmd", required=True)

    sub.add_parser("list-templates", help="List Range category templates")

    sug = sub.add_parser("suggest-template", help="Route SKUs to a template via the category map")
    sug.add_argument("--skus", required=True,
                     help="Space-separated SKU list; an item may be SKU=EAN for barcode fallback")

    init = sub.add_parser("init-batch", help="Init batch state + CSV header")
    init.add_argument("--category", required=True, help="Category leaf or breadcrumb substring")
    init.add_argument("--skus", required=True, help="Space-separated SKU list")
    init.add_argument("--batch-name", default=None)

    gat = sub.add_parser("gather", help="Gather context for one SKU")
    gat.add_argument("--batch-state", required=True)
    gat.add_argument("--sku", required=True)
    gat.add_argument("--ean", default=None,
                     help="EAN/barcode fallback when the SKU isn't a Shopify SKU (SKU-mismatch)")

    wrt = sub.add_parser("write-row", help="Validate + append one row")
    wrt.add_argument("--batch-state", required=True)
    wrt.add_argument("--sku", required=True)
    wrt.add_argument("--row-json", required=True)

    fin = sub.add_parser("finalize", help="Print batch summary")
    fin.add_argument("--batch-state", required=True)

    args = parser.parse_args(argv)
    if args.cmd == "list-templates":
        return _cmd_list_templates()
    if args.cmd == "suggest-template":
        skus = args.skus.split()
        if not skus:
            parser.error("--skus is empty")
        return _cmd_suggest_template(skus)
    if args.cmd == "init-batch":
        skus = args.skus.split()
        if not skus:
            parser.error("--skus is empty")
        return _cmd_init_batch(args.category, skus, args.batch_name)
    if args.cmd == "gather":
        return _cmd_gather(args.batch_state, args.sku, args.ean)
    if args.cmd == "write-row":
        return _cmd_write_row(args.batch_state, args.sku, args.row_json)
    if args.cmd == "finalize":
        return _cmd_finalize(args.batch_state)
    parser.error(f"Unknown command: {args.cmd}")
    return 2


if __name__ == "__main__":
    sys.exit(main())
